import os
import datetime
import sqlite3
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
from uuid import uuid4

from flask import Flask, jsonify, render_template, request, send_from_directory, session, redirect, url_for, g, has_app_context
from functools import wraps
import csv
from exportar_petlove import consultar_lista, salvar_excel, buscar_romaneio_por_pedido, buscar_dados_pedido_tms

BASE_DADOS_DIR = Path(os.getenv("BASE_DADOS_DIR", Path(__file__).parent / "Base de dados"))
DB_PATH = Path(os.getenv("DATABASE_PATH", Path(__file__).parent / "petlove.db"))

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "petlove_jm_secret_key_2026_fixed_session_key")
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(days=30)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario' not in session:
            if request.path.startswith('/api/') or request.is_json:
                return jsonify(erro="Sessão expirada. Por favor, faça login novamente.", redirecionar="/login"), 401
            return redirect(url_for('login', proxima=request.path))
        return f(*args, **kwargs)
    return decorated_function


def buscar_lista_por_pedido(pedido):
    base_dir = BASE_DADOS_DIR
    if not base_dir.exists():
        return None
    
    # Busca em arquivos CSV recursivamente (encontra na pasta Relatorio TMS)
    for arquivo in base_dir.rglob("*.csv"):
        with open(arquivo, newline='', encoding='utf-8-sig', errors='ignore') as f:
            leitor = csv.DictReader(f, delimiter=';')
            # Se não encontrar a coluna, tenta vírgula
            if not leitor.fieldnames or 'Pedido' not in leitor.fieldnames:
                f.seek(0)
                leitor = csv.DictReader(f, delimiter=',')
            if leitor.fieldnames and 'Pedido' in leitor.fieldnames:
                for linha in leitor:
                    if str(linha.get("Pedido", "")).strip() == pedido:
                        return str(linha.get("Lista_Entrega", "")).strip()
    
    # Busca em arquivos XLSX recursivamente
    try:
        from openpyxl import load_workbook
        for arquivo in base_dir.rglob("*.xlsx"):
            wb = load_workbook(arquivo, data_only=True, read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if idx == 0:
                        headers = [str(cell).strip() if cell else "" for cell in row]
                        continue
                    if 'Pedido' in headers and 'Lista_Entrega' in headers:
                        idx_pedido = headers.index('Pedido')
                        idx_lista = headers.index('Lista_Entrega')
                        if str(row[idx_pedido]).strip() == pedido:
                            return str(row[idx_lista]).strip()
    except Exception:
        pass

    return None

EXPORT_DIR = Path(__file__).parent / "exports"
EXPORT_DIR.mkdir(exist_ok=True)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        senha = request.form.get("senha", "").strip()
        
        csv_path = BASE_DADOS_DIR / "usuarios.csv"
        if csv_path.exists():
            # Tenta utf-8-sig primeiro (lida com BOM do Excel/Windows) e depois utf-8
            for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
                try:
                    with open(csv_path, newline='', encoding=encoding, errors='ignore') as f:
                        leitor = csv.DictReader(f)
                        for user in leitor:
                            # Limpa chaves e valores (remove espaços e BOM residuais)
                            user_clean = {}
                            for k, v in user.items():
                                if k is None:
                                    continue
                                k_clean = k.strip().lstrip('\ufeff').lower()
                                user_clean[k_clean] = str(v).strip() if v else ""
                            
                            row_email = (user_clean.get("email") or "").lower()
                            row_senha = user_clean.get("senha") or ""
                            
                            if not row_email:
                                continue
                            
                            is_match = (row_email == email)
                            
                            if is_match and row_senha == senha:
                                session.permanent = True
                                session['usuario'] = row_email
                                session['nome'] = (user_clean.get("nome") or email.split("@")[0].title()).strip()
                                raw_perfil = (user_clean.get("perfil") or "Usuario").strip()
                                # Normaliza: se for 'admin' (qualquer case), salva como 'Admin'
                                if raw_perfil.lower() == "admin":
                                    raw_perfil = "Admin"
                                session['perfil'] = raw_perfil
                                return redirect(request.args.get("proxima") or url_for('index'))
                    break  # Se leu sem erro, para de tentar outros encodings
                except Exception:
                    continue
                        
        return render_template("login.html", erro="Credenciais inválidas")
        
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.post("/api/solicitar-acesso")
def solicitar_acesso():
    import random
    import string
    dados = request.get_json(silent=True) or {}
    email = dados.get("email", "").strip().lower()
    
    if not email or "@" not in email:
        return jsonify(erro="E-mail inválido."), 400
        
    csv_path = BASE_DADOS_DIR / "usuarios.csv"
    campos = ["Email", "Senha", "Perfil", "Nome"]
    linhas_existentes = []
    
    if csv_path.exists():
        with open(csv_path, newline='', encoding='utf-8-sig', errors='ignore') as f:
            leitor = csv.DictReader(f)
            for row in leitor:
                row_clean = {}
                for k, v in row.items():
                    if k:
                        row_clean[k.strip().lstrip('\ufeff')] = str(v).strip() if v else ""
                
                row_email = (row_clean.get("Email") or row_clean.get("email") or "").lower()
                if row_email == email:
                    return jsonify(erro="Este e-mail já possui cadastro."), 400
                linhas_existentes.append(row_clean)
                    
    # Gera senha de 6 caracteres aleatórios
    senha_gerada = ''.join(random.choices(string.ascii_letters + string.digits, k=6))
    
    # Extrai o nome a partir do email (antes do @)
    nome = email.split("@")[0].replace(".", " ").title()

    linhas_existentes.append({
        "Email": email,
        "Senha": senha_gerada,
        "Perfil": "Usuario",
        "Nome": nome
    })

    with open(csv_path, mode='w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=campos, extrasaction='ignore')
        writer.writeheader()
        for row in linhas_existentes:
            writer.writerow(row)

    return jsonify(sucesso=True, senha=senha_gerada)



@app.get("/")
@login_required
def index():
    return render_template("index.html", login_padrao=os.getenv("TMSLOG_LOGIN", ""), usuario=session.get('nome'), perfil=session.get('perfil'))


@app.post("/consultar")
def consultar():
    dados = request.get_json(silent=True) or {}
    numero_pedido = str(dados.get("pedido", "")).strip()
    numero_lista = str(dados.get("numero", "")).strip()
    login = str(dados.get("login", "")).strip()
    senha = str(dados.get("senha", "")).strip()
    if not login or not senha:
        return jsonify(erro="Preencha o login e senha corretamente."), 400
    if not numero_pedido and not numero_lista:
        return jsonify(erro="Preencha o Número do Pedido OU o Número da Lista."), 400
    try:
        if numero_pedido:
            lista_entrega = buscar_lista_por_pedido(numero_pedido)
            if not lista_entrega:
                lista_entrega = buscar_romaneio_por_pedido(numero_pedido, login, senha)
            if not lista_entrega:
                return jsonify(erro=f"Pedido {numero_pedido} não encontrado na Base de dados nem no TMS."), 404
        else:
            lista_entrega = numero_lista

        tabelas = consultar_lista(lista_entrega, login, senha, pedido_alvo=numero_pedido)
        nome = f"lista_{lista_entrega}_{uuid4().hex[:8]}.xlsx"
        nomes = ["Resumo da carga", "Relação da carga"]
        salvar_excel(tabelas, EXPORT_DIR / nome, nomes)
        return jsonify(
            numero=numero_pedido or numero_lista,
            lista=lista_entrega,
            tabelas=tabelas,
            nomes_tabelas=nomes,
            download=f"/baixar/{nome}",
        )
    except Exception as exc:
        mensagem = str(exc)
        if "usuário ou senha inválido" in mensagem.lower():
            mensagem += " Verifique maiúsculas e minúsculas; a senha fornecida inicialmente possui 6 caracteres."
        return jsonify(erro=mensagem), 502


@app.get("/baixar/<path:nome>")
def baixar(nome):
    return send_from_directory(EXPORT_DIR, nome, as_attachment=True)


@app.get("/fluxo")
@login_required
def fluxo():
    return render_template("fluxo.html", login_padrao=os.getenv("TMSLOG_LOGIN", ""), usuario=session.get('nome'), perfil=session.get('perfil'))

# sqlite3 já importado no topo do arquivo

def get_db_connection():
    db_path = DB_PATH
    conn = sqlite3.connect(db_path, timeout=60.0)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA busy_timeout=60000;")
    except Exception:
        pass
    return conn

def get_db():
    if not has_app_context():
        return get_db_connection()
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = get_db_connection()
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

# Criar tabela de recuperação de senha se não existir
try:
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS recuperacao_senha (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT,
            token TEXT,
            senha_provisoria TEXT,
            data_criacao TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS descontos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pedido_id TEXT UNIQUE,
            data_reclamacao TEXT,
            data_carregamento TEXT,
            motorista TEXT,
            regional TEXT,
            rota TEXT,
            placa TEXT,
            valor_desconto TEXT,
            origem_arquivo TEXT,
            data_importacao TEXT,
            observacao TEXT
        )
    ''')
    conn.commit()
    conn.close()
except Exception as e:
    print(f"Erro ao inicializar tabelas SQLite: {e}")

def atualizar_senha_csv(email, nova_senha):
    csv_path = BASE_DADOS_DIR / "usuarios.csv"
    if not csv_path.exists():
        return False
    
    linhas = []
    atualizado = False
    campos = ["Email", "Senha", "Perfil", "Nome"]
    target_email = str(email or "").strip().lower()
    
    with open(csv_path, mode='r', newline='', encoding='utf-8-sig', errors='ignore') as f:
        leitor = csv.DictReader(f)
        for row in leitor:
            row_clean = {}
            for k, v in row.items():
                if k:
                    k_clean = k.strip().lstrip('\ufeff')
                    row_clean[k_clean] = str(v).strip() if v else ""
            
            row_email = (row_clean.get("Email") or row_clean.get("email") or "").lower()
            if row_email == target_email:
                row_clean["Senha"] = nova_senha
                atualizado = True
            linhas.append(row_clean)
            
    if atualizado:
        with open(csv_path, mode='w', newline='', encoding='utf-8') as f:
            escritor = csv.DictWriter(f, fieldnames=campos, extrasaction='ignore')
            escritor.writeheader()
            for row in linhas:
                escritor.writerow(row)
        return True
    return False


# smtplib e email.mime já importados no topo do arquivo

def enviar_email_recuperacao(email, senha_provisoria, link_confirmacao):
    smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
    try:
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
    except:
        smtp_port = 587
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_password = os.getenv("SMTP_PASSWORD", "")
    smtp_from = os.getenv("SMTP_FROM", smtp_user or "noreply@jmdistribuicao.com.br")

    assunto = "Recuperacao de Senha - Portal Petlove JM"
    corpo = f"""Olá,

Foi solicitada a recuperação de senha para sua conta no Portal Petlove JM.

Sua senha temporária gerada é: {senha_provisoria}

Para confirmar a recuperação de senha e redefinir para sua senha desejada, clique no link abaixo:
{link_confirmacao}

Atenção: Este link é obrigatório para ativar seu acesso e definir sua senha definitiva.

Se você não solicitou esta recuperação, por favor desconsidere este e-mail.

Atenciosamente,
Equipe JM Distribuição
"""

    if not smtp_user or not smtp_password:
        print("\n=== [SIMULACAO DE ENVIO DE E-MAIL] ===")
        print(f"Para: {email}")
        print(f"De: {smtp_from}")
        print(f"Assunto: {assunto}")
        print(f"Conteúdo:\n{corpo}")
        print("======================================\n")
        return True, "Simulado no console (sem credenciais SMTP configuradas)."

    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_from
        msg['To'] = email
        msg['Subject'] = assunto
        msg.attach(MIMEText(corpo, 'plain', 'utf-8'))

        server = smtplib.SMTP(smtp_server, smtp_port, timeout=10)
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)
        server.quit()
        return True, "E-mail enviado com sucesso."
    except Exception as e:
        print(f"Erro ao enviar e-mail real: {e}")
        print("\n=== [FALLBACK - SIMULACAO DE E-MAIL] ===")
        print(f"Para: {email}")
        print(f"Conteúdo:\n{corpo}")
        print("========================================\n")
        return False, str(e)

@app.before_request
def verificar_forcar_redefinicao():
    if session.get('forcar_redefinicao'):
        # Rotas permitidas quando forçando redefinição
        rotas_permitidas = ['definir_senha', 'logout', 'static']
        if request.endpoint and request.endpoint not in rotas_permitidas:
            return redirect(url_for('definir_senha'))

@app.post("/api/esqueci-senha")
def api_esqueci_senha():
    import random
    import string
    import datetime
    dados = request.get_json(silent=True) or {}
    email = dados.get("email", "").strip()
    
    if not email or "@" not in email:
        return jsonify(erro="E-mail inválido."), 400
        
    csv_path = BASE_DADOS_DIR / "usuarios.csv"
    email_cadastrado = False
    
    # Verifica se e-mail está cadastrado
    if csv_path.exists():
        with open(csv_path, newline='', encoding='utf-8', errors='ignore') as f:
            leitor = csv.DictReader(f)
            for row in leitor:
                if row.get("Email") == email:
                    email_cadastrado = True
                    break
                    
    if not email_cadastrado:
        return jsonify(erro="Este e-mail não está cadastrado no sistema."), 404
        
    # Gera senha de 6 caracteres aleatórios
    senha_provisoria = ''.join(random.choices(string.ascii_letters + string.digits, k=6))
    token = str(uuid4())
    data_criacao = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Salva no SQLite
    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO recuperacao_senha (email, token, senha_provisoria, data_criacao) VALUES (?, ?, ?, ?)",
            (email, token, senha_provisoria, data_criacao)
        )
        conn.commit()
    except Exception as e:
        return jsonify(erro=f"Erro ao salvar solicitação: {str(e)}"), 500
        
    # Cria o link de confirmação
    url_base = request.host_url.rstrip('/')
    link_confirmacao = f"{url_base}/confirmar-recuperacao?email={email}&token={token}"
    
    sucesso, msg = enviar_email_recuperacao(email, senha_provisoria, link_confirmacao)
    
    return jsonify(sucesso=True, mensagem=msg)

@app.route("/confirmar-recuperacao")
def confirmar_recuperacao():
    email = request.args.get("email", "").strip()
    token = request.args.get("token", "").strip()
    
    if not email or not token:
        return render_template("login.html", erro="Link de confirmação inválido ou incompleto.")
        
    try:
        conn = get_db()
        row = conn.execute("SELECT * FROM recuperacao_senha WHERE email = ? AND token = ?", (email, token)).fetchone()
        
        if not row:
            return render_template("login.html", erro="Token de confirmação inválido ou já utilizado.")
            
        senha_provisoria = row["senha_provisoria"]
        
        # Atualiza a senha no CSV para a provisória
        atualizar_senha_csv(email, senha_provisoria)
        
        # Busca perfil e nome no CSV para montar a sessão
        nome = email.split("@")[0].replace(".", " ").title()
        perfil = "Usuario"
        csv_path = BASE_DADOS_DIR / "usuarios.csv"
        if csv_path.exists():
            with open(csv_path, newline='', encoding='utf-8', errors='ignore') as f:
                leitor = csv.DictReader(f)
                for user in leitor:
                    if user.get("Email") == email:
                        nome = user.get("Nome", nome)
                        perfil = user.get("Perfil", perfil)
                        break
                        
        # Deleta a entrada de recuperação
        conn.execute("DELETE FROM recuperacao_senha WHERE email = ? AND token = ?", (email, token))
        conn.commit()
        
        # Loga o usuário temporariamente e obriga a redefinir a senha
        session['usuario'] = email
        session['nome'] = nome
        session['perfil'] = perfil
        session['forcar_redefinicao'] = True
        
        return redirect(url_for('definir_senha'))
        
    except Exception as e:
        return render_template("login.html", erro=f"Erro interno ao confirmar recuperação: {str(e)}")

@app.route("/definir-senha", methods=["GET", "POST"])
def definir_senha():
    if not session.get('forcar_redefinicao'):
        return redirect(url_for('login'))
        
    email = session.get('usuario')
    
    if request.method == "POST":
        nova_senha = request.form.get("nova_senha", "").strip()
        confirmacao = request.form.get("confirmacao", "").strip()
        
        if not nova_senha:
            return render_template("definir_senha.html", erro="A senha não pode ser vazia.")
            
        if nova_senha != confirmacao:
            return render_template("definir_senha.html", erro="As senhas não coincidem.")
            
        # Atualiza a senha no CSV para a definitiva
        if atualizar_senha_csv(email, nova_senha):
            # Limpa flag de redefinição
            session.pop('forcar_redefinicao', None)
            return redirect(url_for('index'))
        else:
            return render_template("definir_senha.html", erro="Erro ao atualizar a senha no banco de dados.")
            
    return render_template("definir_senha.html")

def is_admin():
    user = str(session.get("usuario", "")).strip().lower()
    perfil = str(session.get("perfil", "")).strip().lower()
    if perfil == "admin" or user == "admin@jm.com":
        return True
    if user:
        csv_path = BASE_DADOS_DIR / "usuarios.csv"
        if csv_path.exists():
            try:
                for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
                    try:
                        with open(csv_path, newline='', encoding=enc, errors='ignore') as f:
                            leitor = csv.DictReader(f)
                            for row in leitor:
                                u_clean = {}
                                for k, v in row.items():
                                    if k: u_clean[k.strip().lstrip('\ufeff').lower()] = str(v).strip()
                                u_email = (u_clean.get("email") or "").lower()
                                u_perfil = (u_clean.get("perfil") or "").lower()
                                if u_email == user and u_perfil == "admin":
                                    session['perfil'] = "Admin"
                                    return True
                        break
                    except Exception:
                        continue
            except Exception:
                pass
    return False

@app.context_processor
def inject_is_admin():
    return dict(is_admin=is_admin())

@app.route("/usuarios")
@login_required
def usuarios_dashboard():
    if not is_admin():
        return redirect(url_for('gestao'))
    return render_template("usuarios.html", usuario=session.get('nome'), perfil=session.get('perfil'))

@app.route("/api/usuarios")
@login_required
def api_listar_usuarios():
    if not is_admin():
        return jsonify(erro="Acesso negado. Apenas administradores podem realizar esta ação."), 403
        
    csv_path = BASE_DADOS_DIR / "usuarios.csv"
    if not csv_path.exists():
        return jsonify([])
        
    usuarios = []
    for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            with open(csv_path, newline='', encoding=encoding, errors='ignore') as f:
                leitor = csv.DictReader(f)
                for row in leitor:
                    row_clean = {}
                    for k, v in row.items():
                        if k:
                            k_clean = k.strip().lstrip('\ufeff')
                            row_clean[k_clean] = str(v).strip() if v else ""
                    
                    row_email = row_clean.get("Email") or row_clean.get("email") or ""
                    if not row_email:
                        continue
                        
                    usuarios.append({
                        "Email": row_email,
                        "Nome": row_clean.get("Nome") or row_clean.get("nome") or "",
                        "Perfil": row_clean.get("Perfil") or row_clean.get("perfil") or "Usuario",
                        "Senha": row_clean.get("Senha") or row_clean.get("senha") or ""
                    })
            break
        except Exception:
            continue
    return jsonify(usuarios)

@app.route("/api/usuarios/salvar", methods=["POST"])
@login_required
def api_salvar_usuario():
    if not is_admin():
        return jsonify(erro="Acesso negado. Apenas administradores podem realizar esta ação."), 403
        
    dados = request.get_json(silent=True) or {}
    email = dados.get("email", "").strip().lower()
    nome = dados.get("nome", "").strip()
    perfil = dados.get("perfil", "").strip()
    senha = dados.get("senha", "").strip()
    is_edit = dados.get("is_edit", False)
    
    if not email or not nome or not perfil:
        return jsonify(erro="Preencha os campos obrigatórios (E-mail, Nome e Perfil)."), 400
        
    if perfil not in ["Admin", "Usuario"]:
        return jsonify(erro="Perfil inválido."), 400
        
    csv_path = BASE_DADOS_DIR / "usuarios.csv"
    
    linhas = []
    editado = False
    campos = ["Email", "Senha", "Perfil", "Nome"]
    email_existe = False
    
    if csv_path.exists():
        with open(csv_path, newline='', encoding='utf-8-sig', errors='ignore') as f:
            leitor = csv.DictReader(f)
            for row in leitor:
                row_clean = {}
                for k, v in row.items():
                    if k:
                        k_clean = k.strip().lstrip('\ufeff')
                        row_clean[k_clean] = str(v).strip() if v else ""
                
                row_email = (row_clean.get("Email") or row_clean.get("email") or "").lower()
                if row_email == email:
                    email_existe = True
                    if is_edit:
                        row_clean["Nome"] = nome
                        row_clean["Perfil"] = perfil
                        if senha:
                            row_clean["Senha"] = senha
                        editado = True
                linhas.append(row_clean)
                
    if not is_edit and email_existe:
        return jsonify(erro="Este e-mail de usuário já está cadastrado. Para alterá-lo, use o botão Editar na listagem."), 400
        
    if not editado and is_edit:
        return jsonify(erro="Usuário para edição não encontrado no sistema."), 404
        
    if not editado:
        # Novo usuário. Senha é obrigatória para novos usuários
        if not senha:
            return jsonify(erro="A senha é obrigatória para novos usuários."), 400
        linhas.append({
            "Email": email,
            "Senha": senha,
            "Perfil": perfil,
            "Nome": nome
        })
        
    # Grava de volta no CSV
    try:
        with open(csv_path, mode='w', newline='', encoding='utf-8') as f:
            escritor = csv.DictWriter(f, fieldnames=campos, extrasaction='ignore')
            escritor.writeheader()
            for row in linhas:
                row["Email"] = str(row.get("Email") or row.get("email") or "").strip().lower()
                escritor.writerow(row)
        return jsonify(sucesso=True)
    except Exception as e:
        return jsonify(erro=f"Erro ao salvar usuário: {str(e)}"), 500

@app.route("/api/usuarios/deletar", methods=["POST"])
@login_required
def api_deletar_usuario():
    if not is_admin():
        return jsonify(erro="Acesso negado. Apenas administradores podem realizar esta ação."), 403
        
    dados = request.get_json(silent=True) or {}
    email = dados.get("email", "").strip().lower()
    
    if not email:
        return jsonify(erro="E-mail do usuário não informado."), 400
        
    # Impedir que o administrador delete a si mesmo
    if email == str(session.get("usuario")).strip().lower():
        return jsonify(erro="Você não pode excluir sua própria conta."), 400
        
    csv_path = BASE_DADOS_DIR / "usuarios.csv"
    if not csv_path.exists():
        return jsonify(erro="Arquivo de usuários não encontrado."), 404
        
    linhas = []
    deletado = False
    campos = ["Email", "Senha", "Perfil", "Nome"]
    
    with open(csv_path, newline='', encoding='utf-8-sig', errors='ignore') as f:
        leitor = csv.DictReader(f)
        for row in leitor:
            row_clean = {}
            for k, v in row.items():
                if k:
                    k_clean = k.strip().lstrip('\ufeff')
                    row_clean[k_clean] = str(v).strip() if v else ""
            
            row_email = (row_clean.get("Email") or row_clean.get("email") or "").lower()
            if row_email == email:
                deletado = True
                continue
            linhas.append(row_clean)
            
    if not deletado:
        return jsonify(erro="Usuário não encontrado."), 404
        
    try:
        with open(csv_path, mode='w', newline='', encoding='utf-8') as f:
            escritor = csv.DictWriter(f, fieldnames=campos, extrasaction='ignore')
            escritor.writeheader()
            for row in linhas:
                escritor.writerow(row)
        return jsonify(sucesso=True)
    except Exception as e:
        return jsonify(erro=f"Erro ao excluir usuário: {str(e)}"), 500


def sincronizar_pasta_tms():
    db_path = DB_PATH
    tms_dir = BASE_DADOS_DIR / "Relatorio TMS"
    
    if not tms_dir.exists():
        return 0, 0, "Pasta Relatorio TMS não encontrada."
        
    csv_files = list(tms_dir.glob("*.csv"))
    if not csv_files:
        return 0, 0, "Nenhum arquivo CSV encontrado na pasta Relatorio TMS."
        
    conn = get_db()
    cursor = conn.cursor()
    
    # Obter colunas reais da tabela chamados
    cursor.execute("SELECT * FROM chamados LIMIT 1")
    colunas_db = [description[0] for description in cursor.description]
    
    # Obter todos os chamados existentes na base
    cursor.execute("SELECT * FROM chamados")
    rows = cursor.fetchall()
    
    col_id_db = next((c for c in colunas_db if 'ID_do_Pedido' in c or 'ID_Pedido' in c), 'ID_do_Pedido')
    
    chamados_existentes = {}
    for r in rows:
        d = dict(r)
        key_map = {}
        for k in d.keys():
            k_clean = k.replace("_", " ").lower()
            if "id do pedido" in k_clean or "id_pedido" in k_clean:
                key_map["id"] = k
            elif "lista entrega cruzada" in k_clean or "lista_entrega_cruzada" in k_clean:
                key_map["lista"] = k
            elif "motorista" in k_clean:
                key_map["motorista"] = k
            elif "regional" in k_clean:
                key_map["regional"] = k
            elif "data do carregamento" in k_clean or "data_do_carregamento" in k_clean:
                key_map["data_carr"] = k
            elif "rota" in k_clean:
                key_map["rota"] = k
            elif "descricao da divergencia" in k_clean or "descrição da divergência" in k_clean or "descrição_da_divergência" in k_clean:
                key_map["divergencia"] = k
                
        chamados_existentes[str(d.get(key_map.get("id", col_id_db), "")).strip()] = (d, key_map)
        
    total_novos = 0
    total_lidos = 0
    
    for csv_file in csv_files:
        delimiter = ','
        with open(csv_file, newline='', encoding='utf-8-sig', errors='ignore') as f:
            first_line = f.readline()
            if first_line.count(';') > first_line.count(','):
                delimiter = ';'
            f.seek(0)
            
            leitor = csv.DictReader(f, delimiter=delimiter)
            if not leitor.fieldnames:
                continue
                
            for row in leitor:
                total_lidos += 1
                col_pedido = next((c for c in leitor.fieldnames if 'Pedido' in c), 'Pedido')
                pedido_id = str(row.get(col_pedido, "")).strip()
                
                if not pedido_id:
                    continue
                    
                # Preparar valores do TMS
                data_carr_raw = row.get("Data_do_Carregamento_Lista", "")
                data_carr = data_carr_raw.split(" ")[0].strip() if " " in data_carr_raw else data_carr_raw.strip()
                
                motorista_raw = row.get("Motorista_Lista", "")
                motorista = motorista_raw.split(" - ", 1)[1].strip() if " - " in motorista_raw else motorista_raw.strip()
                
                filial_raw = row.get("Filial_Entrega", "")
                regional = filial_raw.strip()
                if "São Paulo" in filial_raw: regional = "JM SP"
                elif "Barueri" in filial_raw: regional = "JM BAR"
                elif "Santos" in filial_raw: regional = "JM SSZ"
                
                lista_entrega = row.get("Lista_Entrega", "").strip()
                ocorrencia = row.get("Ultima_Ocorrencia", "").strip()
                rota = row.get("Rota_Entrega", "").strip()
                placa = row.get("Placa_do_veiculo", row.get("Placa", "")).strip()

                if pedido_id not in chamados_existentes:
                    # Inserir novo cadastro no banco chamado
                    novo_dict = {}
                    for col in colunas_db:
                        c_clean = col.replace("_", " ").lower()
                        if "id do pedido" in c_clean or "id_pedido" in c_clean:
                            novo_dict[col] = pedido_id
                        elif "data do carregamento" in c_clean or "data_do_carregamento" in c_clean:
                            novo_dict[col] = data_carr
                        elif "motorista" in c_clean:
                            novo_dict[col] = motorista
                        elif "regional" in c_clean:
                            novo_dict[col] = regional
                        elif "rota" in c_clean:
                            novo_dict[col] = rota
                        elif "placa" in c_clean:
                            novo_dict[col] = placa
                        elif "descricao da divergencia" in c_clean or "descrição" in c_clean:
                            novo_dict[col] = ocorrencia
                        elif "lista entrega cruzada" in c_clean:
                            novo_dict[col] = lista_entrega
                        elif "status da tratativa" in c_clean or "status_da_tratativa" in c_clean:
                            novo_dict[col] = "Em Andamento"
                        elif "procedencia" in c_clean:
                            novo_dict[col] = "Em Análise"
                        elif "criado por" in c_clean or "criado_por" in c_clean:
                            novo_dict[col] = "Carga TMS"
                        else:
                            novo_dict[col] = ""
                    
                    placeholders = ", ".join(["?"] * len(colunas_db))
                    cols_str = ", ".join([f'"{c}"' for c in colunas_db])
                    sql_ins = f'INSERT INTO chamados ({cols_str}) VALUES ({placeholders})'
                    vals_ins = [novo_dict.get(c, "") for c in colunas_db]
                    cursor.execute(sql_ins, vals_ins)
                    
                    chamados_existentes[pedido_id] = (novo_dict, {})
                    total_novos += 1
                    continue
                    
                # O pedido existe na base! Vamos atualizar campos que estiverem em branco.
                chamado_dict, key_map = chamados_existentes[pedido_id]
                
                updates = {}
                
                col_lista = key_map.get("lista")
                if col_lista and not str(chamado_dict.get(col_lista, "")).strip():
                    updates[col_lista] = lista_entrega
                    
                col_mot = key_map.get("motorista")
                if col_mot and not str(chamado_dict.get(col_mot, "")).strip():
                    updates[col_mot] = motorista
                    
                col_reg = key_map.get("regional")
                if col_reg and not str(chamado_dict.get(col_reg, "")).strip():
                    updates[col_reg] = regional
                    
                col_data = key_map.get("data_carr")
                if col_data and not str(chamado_dict.get(col_data, "")).strip():
                    updates[col_data] = data_carr
                    
                col_rota = key_map.get("rota")
                if col_rota and not str(chamado_dict.get(col_rota, "")).strip():
                    updates[col_rota] = row.get("Rota_Entrega", "").strip()
                    
                col_diverg = key_map.get("divergencia")
                if col_diverg and not str(chamado_dict.get(col_diverg, "")).strip():
                    updates[col_diverg] = ocorrencia
                    
                if updates:
                    set_clause = ", ".join([f'"{k}" = ?' for k in updates.keys()])
                    sql = f'UPDATE chamados SET {set_clause} WHERE "{col_id_db}" = ?'
                    params = list(updates.values()) + [pedido_id]
                    cursor.execute(sql, params)
                    
                    total_novos += 1
                    for k, v in updates.items():
                        chamado_dict[k] = v
                        
    conn.commit()
    return total_novos, total_lidos, None

@app.route("/api/tms/sincronizar", methods=["POST"])
@login_required
def api_sincronizar_tms():
    if not is_admin():
        return jsonify(erro="Acesso negado. Apenas administradores podem realizar esta ação."), 403
        
    try:
        total_novos, total_lidos, erro = sincronizar_pasta_tms()
        if erro:
            return jsonify(erro=erro), 400
        return jsonify(sucesso=True, total_novos=total_novos, total_lidos=total_lidos)
    except Exception as e:
        return jsonify(erro=str(e)), 500

def obter_inconsistencias():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM chamados")
    rows = cursor.fetchall()
    
    colunas_db = [description[0] for description in cursor.description]
    col_id_db = next((c for c in colunas_db if 'ID_do_Pedido' in c or 'ID_Pedido' in c), 'ID_do_Pedido')
    
    inconsistencias = []
    for r in rows:
        d = dict(r)
        
        # Mapeando chaves do dicionário para compatibilidade de nomes com acentos e espaços
        d_clean = {}
        for k, v in d.items():
            k_clean = k.replace("_", " ").lower().strip()
            d_clean[k_clean] = str(v or "").strip()
            
        pid = d_clean.get("id do pedido") or d_clean.get("id_pedido") or ""
        data_carr = d_clean.get("data do carregamento") or d_clean.get("data_do_carregamento") or ""
        motorista = d_clean.get("motorista") or ""
        placa = d_clean.get("placa do veiculo") or d_clean.get("placa_do_veiculo") or d_clean.get("placa do veículo") or ""
        criado = d.get("Criado_por") or d.get("Criado por") or ""
        regional = d_clean.get("regional 2") or d_clean.get("regional") or ""
        rota = d_clean.get("rota") or ""
        
        erros = []
        
        # 1. ID Pedido (deve ser numérico com 9 dígitos)
        if not pid:
            erros.append("ID do Pedido está em branco.")
        elif not pid.isdigit() or len(pid) != 9:
            erros.append(f"ID do Pedido inválido (deve ter 9 dígitos numéricos: '{pid}').")
            
        # 2. Data do Carregamento (deve ser DD/MM/AAAA)
        if not data_carr:
            erros.append("Data do Carregamento está em branco.")
        elif len(data_carr) != 10 or '/' not in data_carr:
            erros.append(f"Data do Carregamento com formato inválido (esperado DD/MM/AAAA, recebido: '{data_carr}').")
            
        # 3. Regional
        if not regional:
            erros.append("Regional está em branco.")
            
        # 4. Rota
        if not rota:
            erros.append("Rota está em branco.")
            
        # 5. Motorista
        if not motorista:
            erros.append("Motorista está em branco.")
            
        # 6. Placa (se preenchida, deve ter formato de placa brasileira ABC1234 ou ABC1D23)
        if placa:
            placa_clean = placa.replace("-", "").strip()
            if len(placa_clean) != 7 or not placa_clean[:3].isalpha() or not (placa_clean[3:].isalnum()):
                erros.append(f"Placa do veículo inválida (esperado ABC-1234 ou ABC1D23, recebido: '{placa}').")
                
        if erros:
            inconsistencias.append({
                "id": pid,
                "criado_por": criado or "Não Informado",
                "motorista": motorista or "Não Informado",
                "data_carregamento": data_carr or "Não Informada",
                "erros": erros,
                "dados": {
                    "Regional": regional,
                    "Rota": rota,
                    "Placa": placa,
                    "Valor": d_clean.get("valor", "")
                }
            })
            
    return inconsistencias

@app.route("/inconsistencias")
@login_required
def inconsistencias_dashboard():
    if not is_admin():
        return redirect(url_for('gestao'))
    return render_template("inconsistencias.html", usuario=session.get('nome'), perfil=session.get('perfil'))

@app.route("/api/inconsistencias")
@login_required
def api_inconsistencias():
    if not is_admin():
        return jsonify(erro="Acesso negado. Apenas administradores podem realizar esta ação."), 403
    try:
        dados = obter_inconsistencias()
        return jsonify(dados)
    except Exception as e:
        return jsonify(erro=str(e)), 500

def dict_from_row(row):
    d = dict(row)
    # Renomear as chaves de volta para o formato esperado pelo frontend
    mapping = {
        "Descri\xe7\xe3o_da_Reclama\xe7\xe3o": "Descri\xe7\xe3o da Reclama\xe7\xe3o",
        "Data_do_Carregamento": "Data do Carregamento",
        "ID_do_Pedido": "ID_do_Pedido",
        "Motorista": "Motorista",
        "Regional_2": "Regional_2",
        "Rota": "Rota",
        "Placa_do_ve\xedculo": "Placa do ve\xedculo",
        "Valor": "Valor",
        "Justificativa": "Justificativa",
        "Descri\xe7\xe3o_da_Diverg\xeancia": "Descri\xe7\xe3o da Diverg\xeancia",
        "Proced\xeancia": "Proced\xeancia",
        "Responsavel": "Responsavel",
        "Tratativa": "Tratativa",
        "Criado_por": "Criado por"
    }
    # Decodificar chaves corrompidas se necessário
    new_d = {}
    for k, v in d.items():
        # Vamos tentar associar pelo prefixo ou simplesmente substituir os underscores
        new_key = k.replace("_", " ")
        if "ID do Pedido" in new_key: new_key = "ID_do_Pedido"
        if "Regional 2" in new_key: new_key = "Regional_2"
        if "Criado por" in new_key: new_key = "Criado por"
        
        val = v
        if isinstance(v, str):
            val_clean = v.strip()
            if "proced" in new_key.lower():
                val_lower = val_clean.lower()
                if val_lower in ["procedente", "procédente", "procedente"]:
                    val = "Procedente"
                elif val_lower in ["não procedente", "nao procedente", "improcedente", "improcedente"]:
                    val = "Não Procedente"
                elif val_lower == "em analise" or val_lower == "em análise":
                    val = "Em Análise"
                else:
                    val = val_clean
            else:
                val = val_clean
        new_d[new_key] = val
    return new_d

@app.get("/api/fluxo/buscar/<pedido>")
@login_required
def fluxo_buscar(pedido):
    try:
        conn = get_db()
        row = conn.execute("SELECT * FROM chamados WHERE ID_do_Pedido = ?", (str(pedido).strip(),)).fetchone()
        
        if row:
            linha = dict_from_row(row)
            try:
                # Se a coluna existir e tiver valor no banco, usa ele. Caso contrário busca no CSV.
                colunas = row.keys() if hasattr(row, 'keys') else []
                if 'Lista_Entrega_Cruzada' in colunas and row['Lista_Entrega_Cruzada'] and str(row['Lista_Entrega_Cruzada']).strip():
                    linha['Lista_Entrega_Cruzada'] = str(row['Lista_Entrega_Cruzada']).strip()
                else:
                    linha['Lista_Entrega_Cruzada'] = buscar_lista_por_pedido(pedido) or "Não encontrada"
            except Exception as e:
                # Tenta chamar a função local buscando no CSV
                try:
                    linha['Lista_Entrega_Cruzada'] = buscar_lista_por_pedido(pedido) or "Não encontrada"
                except:
                    linha['Lista_Entrega_Cruzada'] = "Erro"
            
            # Adicionar histórico de modificações do pedido
            try:
                hist_rows = conn.execute("SELECT data_hora, usuario, campo, valor_antigo, valor_novo FROM historico_chamados WHERE pedido_id = ? ORDER BY id DESC", (str(pedido).strip(),)).fetchall()
                linha['Historico'] = [dict(r) for r in hist_rows]
            except Exception as e:
                linha['Historico'] = []
                
            return jsonify(linha)
            
        return jsonify(erro="Pedido não encontrado no fluxo."), 404
    except Exception as e:
        return jsonify(erro=str(e)), 500

@app.post("/api/fluxo/atualizar")
@login_required
def fluxo_atualizar():
    dados = request.get_json(silent=True) or {}
    pedido = str(dados.get("pedido", "")).strip()
    usuario = session.get('nome', 'Usuário')
    
    if not pedido:
        return jsonify(erro="Pedido não informado."), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 1. Buscar valores antigos para o histórico
        cursor.execute("SELECT * FROM chamados WHERE ID_do_Pedido = ?", (pedido,))
        row_antiga = cursor.fetchone()
        if not row_antiga:
            return jsonify(erro="Pedido não encontrado no fluxo."), 404
            
        col_names = [description[0] for description in cursor.description]
        col_justificativa = next((c for c in col_names if 'Justificativa' in c), 'Justificativa')
        col_procedencia = next((c for c in col_names if 'Proced' in c), 'Procedência')
        col_tratativa = next((c for c in col_names if 'Tratativa' in c), 'Tratativa')
        col_responsavel = next((c for c in col_names if 'Responsavel' in c), 'Responsavel')
        col_divergencia = next((c for c in col_names if 'Diverg' in c), 'Descrição_da_Divergência')
        col_valor = next((c for c in col_names if 'Valor' in c), 'Valor')
        col_status = next((c for c in col_names if 'Status' in c), 'Status_da_Tratativa')
        
        valores_antigos = dict(row_antiga)
        
        nova_justif = dados.get("justificativa", "")
        nova_proc = dados.get("procedencia", "")
        nova_trat = dados.get("tratativa", "")
        novo_resp = dados.get("responsavel", "")
        nova_diverg = dados.get("divergencia", "")
        novo_valor = dados.get("valor", "")
        
        was_finalizado = valores_antigos.get(col_status) and str(valores_antigos.get(col_status)).strip().lower() == 'finalizado'
        
        proc_clean = nova_proc.strip().lower() if nova_proc else ""
        trat_clean = nova_trat.strip() if nova_trat else ""
        resp_clean = novo_resp.strip() if novo_resp else ""
        
        # Determinar se o chamado deve estar finalizado
        novo_status = dados.get("status")
        if not novo_status or novo_status not in ["Em Andamento", "Finalizado"]:
            tem_procedencia_valida = proc_clean and proc_clean not in ["em analise", "em análise"]
            if resp_clean or trat_clean or tem_procedencia_valida:
                novo_status = "Finalizado"
            elif was_finalizado:
                # PROTEÇÃO: Se já estava Finalizado, mantém Finalizado por segurança
                # Impede regressão acidental de status
                novo_status = "Finalizado"
            else:
                novo_status = "Em Andamento"
            
        # 2. Executar o UPDATE
        sql = f"""
            UPDATE chamados 
            SET "{col_justificativa}" = ?,
                "{col_procedencia}" = ?,
                "{col_tratativa}" = ?,
                "{col_responsavel}" = ?,
                "{col_divergencia}" = ?,
                "{col_valor}" = ?,
                "{col_status}" = ?
            WHERE ID_do_Pedido = ?
        """
        
        cursor.execute(sql, (
            nova_justif,
            nova_proc,
            nova_trat,
            novo_resp,
            nova_diverg,
            novo_valor,
            novo_status,
            pedido
        ))
        
        # 3. Registrar modificações no histórico
        # datetime já importado no topo do arquivo
        data_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        
        def log_mudanca(campo, valor_ant, valor_nov):
            val_ant_str = str(valor_ant or "").strip()
            val_nov_str = str(valor_nov or "").strip()
            if val_ant_str != val_nov_str:
                cursor.execute("""
                    INSERT INTO historico_chamados (pedido_id, data_hora, usuario, campo, valor_antigo, valor_novo)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (pedido, data_hora, usuario, campo, val_ant_str if val_ant_str else "Vazio", val_nov_str if val_nov_str else "Vazio"))

        log_mudanca("Justificativa", valores_antigos.get(col_justificativa), nova_justif)
        log_mudanca("Procedência", valores_antigos.get(col_procedencia), nova_proc)
        log_mudanca("Tratativa", valores_antigos.get(col_tratativa), nova_trat)
        log_mudanca("Responsável", valores_antigos.get(col_responsavel), novo_resp)
        log_mudanca("Divergência", valores_antigos.get(col_divergencia), nova_diverg)
        log_mudanca("Valor", valores_antigos.get(col_valor), novo_valor)
        log_mudanca("Status da Tratativa", valores_antigos.get(col_status), novo_status)
        
        conn.commit()
        return jsonify(sucesso=True)
    except Exception as e:
        return jsonify(erro=str(e)), 500

@app.post("/api/fluxo/atualizar_romaneio")
@login_required
def fluxo_atualizar_romaneio():
    dados = request.get_json(silent=True) or {}
    pedido = str(dados.get("pedido", "")).strip()
    novo_romaneio = str(dados.get("romaneio", "")).strip()
    usuario = session.get('nome', 'Usuário')
    
    if not pedido:
        return jsonify(erro="Pedido não informado."), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM chamados WHERE ID_do_Pedido = ?", (pedido,))
        row = cursor.fetchone()
        if not row:
            return jsonify(erro="Pedido não encontrado."), 404
            
        col_names = [description[0] for description in cursor.description]
        col_romaneio = next((c for c in col_names if 'Lista_Entrega_Cruzada' in c), 'Lista_Entrega_Cruzada')
        
        valores_antigos = dict(row)
        valor_antigo = valores_antigos.get(col_romaneio) or ""
        
        if str(valor_antigo).strip() == novo_romaneio:
            return jsonify(sucesso=True)
            
        # Executa atualização
        cursor.execute(f'UPDATE chamados SET "{col_romaneio}" = ? WHERE ID_do_Pedido = ?', (novo_romaneio, pedido))
        
        # Registrar no histórico
        # datetime já importado no topo do arquivo
        data_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        cursor.execute("""
            INSERT INTO historico_chamados (pedido_id, data_hora, usuario, campo, valor_antigo, valor_novo)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (pedido, data_hora, usuario, "Nº Romaneio", valor_antigo if valor_antigo else "Vazio/Erro", novo_romaneio))
        
        conn.commit()
        return jsonify(sucesso=True)
    except Exception as e:
        return jsonify(erro=str(e)), 500

@app.post("/api/fluxo/atualizar_status")
@login_required
def fluxo_atualizar_status():
    dados = request.get_json(silent=True) or {}
    pedido = str(dados.get("pedido", "")).strip()
    novo_status = str(dados.get("status", "")).strip()
    usuario = session.get('nome', 'Usuário')
    
    if not pedido:
        return jsonify(erro="Pedido não informado."), 400
    if novo_status not in ["Em Andamento", "Finalizado"]:
        return jsonify(erro="Status inválido."), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM chamados WHERE ID_do_Pedido = ?", (pedido,))
        row = cursor.fetchone()
        if not row:
            return jsonify(erro="Pedido não encontrado."), 404
            
        col_names = [description[0] for description in cursor.description]
        col_status = next((c for c in col_names if 'Status' in c), 'Status_da_Tratativa')
        
        valores_antigos = dict(row)
        valor_antigo = valores_antigos.get(col_status) or ""
        
        if str(valor_antigo).strip() == novo_status:
            return jsonify(sucesso=True)
            
        # Executa atualização
        cursor.execute(f'UPDATE chamados SET "{col_status}" = ? WHERE ID_do_Pedido = ?', (novo_status, pedido))
        
        # Registrar no histórico
        data_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        cursor.execute("""
            INSERT INTO historico_chamados (pedido_id, data_hora, usuario, campo, valor_antigo, valor_novo)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (pedido, data_hora, usuario, "Status da Tratativa", valor_antigo if valor_antigo else "Vazio", novo_status))
        
        conn.commit()
        return jsonify(sucesso=True)
    except Exception as e:
        return jsonify(erro=str(e)), 500

@app.post("/api/fluxo/atualizar_justificativa")
@login_required
def fluxo_atualizar_justificativa():
    dados = request.get_json(silent=True) or {}
    pedido = str(dados.get("pedido", "")).strip()
    nova_justif = str(dados.get("justificativa", "")).strip()
    usuario = session.get('nome', 'Usuário')
    
    if not pedido:
        return jsonify(erro="Pedido não informado."), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM chamados WHERE ID_do_Pedido = ?", (pedido,))
        row = cursor.fetchone()
        if not row:
            return jsonify(erro="Pedido não encontrado."), 404
            
        col_names = [description[0] for description in cursor.description]
        col_justif = next((c for c in col_names if 'Justificativa' in c), 'Justificativa')
        
        valores_antigos = dict(row)
        valor_antigo = valores_antigos.get(col_justif) or ""
        
        if str(valor_antigo).strip() == nova_justif:
            return jsonify(sucesso=True)
            
        cursor.execute(f'UPDATE chamados SET "{col_justif}" = ? WHERE ID_do_Pedido = ?', (nova_justif, pedido))
        
        data_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        cursor.execute("""
            INSERT INTO historico_chamados (pedido_id, data_hora, usuario, campo, valor_antigo, valor_novo)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (pedido, data_hora, usuario, "Justificativa", valor_antigo if valor_antigo else "Vazio", nova_justif))
        
        conn.commit()
        return jsonify(sucesso=True)
    except Exception as e:
        return jsonify(erro=str(e)), 500

@app.post("/api/fluxo/atualizar_tratativa")
@login_required
def fluxo_atualizar_tratativa():
    dados = request.get_json(silent=True) or {}
    pedido = str(dados.get("pedido", "")).strip()
    nova_tratativa = str(dados.get("tratativa", "")).strip()
    usuario = session.get('nome', 'Usuário')
    
    if not pedido:
        return jsonify(erro="Pedido não informado."), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM chamados WHERE ID_do_Pedido = ?", (pedido,))
        row = cursor.fetchone()
        if not row:
            return jsonify(erro="Pedido não encontrado."), 404
            
        col_names = [description[0] for description in cursor.description]
        col_trat = next((c for c in col_names if 'Tratativa' in c), 'Tratativa')
        
        valores_antigos = dict(row)
        valor_antigo = valores_antigos.get(col_trat) or ""
        
        if str(valor_antigo).strip() == nova_tratativa:
            return jsonify(sucesso=True)
            
        cursor.execute(f'UPDATE chamados SET "{col_trat}" = ? WHERE ID_do_Pedido = ?', (nova_tratativa, pedido))
        
        data_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        cursor.execute("""
            INSERT INTO historico_chamados (pedido_id, data_hora, usuario, campo, valor_antigo, valor_novo)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (pedido, data_hora, usuario, "Tratativa", valor_antigo if valor_antigo else "Vazio", nova_tratativa))
        
        conn.commit()
        return jsonify(sucesso=True)
    except Exception as e:
        return jsonify(erro=str(e)), 500

@app.post("/api/fluxo/atualizar_procedencia")
@login_required
def fluxo_atualizar_procedencia():
    dados = request.get_json(silent=True) or {}
    pedido = str(dados.get("pedido", "")).strip()
    nova_procedencia = str(dados.get("procedencia", "")).strip()
    usuario = session.get('nome', 'Usuário')
    
    if not pedido:
        return jsonify(erro="Pedido não informado."), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM chamados WHERE ID_do_Pedido = ?", (pedido,))
        row = cursor.fetchone()
        if not row:
            return jsonify(erro="Pedido não encontrado."), 404
            
        col_names = [description[0] for description in cursor.description]
        col_proc = next((c for c in col_names if 'Proced' in c), 'Procedência')
        
        valores_antigos = dict(row)
        valor_antigo = valores_antigos.get(col_proc) or ""
        
        if str(valor_antigo).strip() == nova_procedencia:
            return jsonify(sucesso=True)
            
        cursor.execute(f'UPDATE chamados SET "{col_proc}" = ? WHERE ID_do_Pedido = ?', (nova_procedencia, pedido))
        
        data_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        cursor.execute("""
            INSERT INTO historico_chamados (pedido_id, data_hora, usuario, campo, valor_antigo, valor_novo)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (pedido, data_hora, usuario, "Procedência", valor_antigo if valor_antigo else "Vazio", nova_procedencia))
        
        conn.commit()
        return jsonify(sucesso=True)
    except Exception as e:
        return jsonify(erro=str(e)), 500

@app.post("/api/fluxo/deletar")
@login_required
def fluxo_deletar():
    if not is_admin():
        return jsonify(erro="Apenas administradores podem excluir chamados."), 403

    dados = request.get_json(silent=True) or {}
    pedido = str(dados.get("pedido", "")).strip()
    if not pedido:
        return jsonify(erro="Pedido não informado."), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Deleta de chamados
        cursor.execute("DELETE FROM chamados WHERE ID_do_Pedido = ?", (pedido,))
        # Deleta de historico
        cursor.execute("DELETE FROM historico_chamados WHERE pedido_id = ?", (pedido,))
        
        conn.commit()
        return jsonify(sucesso=True)
    except Exception as e:
        return jsonify(erro=str(e)), 500

@app.post("/api/fluxo/atualizar_campos_admin")
@login_required
def fluxo_atualizar_campos_admin():
    if not is_admin():
        return jsonify(erro="Apenas administradores podem editar estes campos."), 403
        
    dados = request.get_json(silent=True) or {}
    pedido = str(dados.get("pedido", "")).strip()
    if not pedido:
        return jsonify(erro="Pedido não informado."), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Obter colunas reais do banco
        cursor.execute("SELECT * FROM chamados WHERE ID_do_Pedido = ?", (pedido,))
        row = cursor.fetchone()
        if not row:
            return jsonify(erro="Pedido não encontrado."), 404
            
        col_names = [description[0] for description in cursor.description]
        
        col_data_rota = next((c for c in col_names if 'Carregamento' in c or 'Data_do' in c), 'Data_do_Carregamento')
        col_data_rec = next((c for c in col_names if 'Reclama' in c or 'Reclamacao' in c), 'Descrição_da_Reclamação')
        col_motorista = next((c for c in col_names if 'Motorista' in c), 'Motorista')
        col_placa = next((c for c in col_names if 'Placa' in c), 'Placa_do_veículo')
        col_rota = next((c for c in col_names if 'Rota' in c), 'Rota')
        col_regional = next((c for c in col_names if 'Regional' in c), 'Regional_2')
        col_valor = next((c for c in col_names if 'Valor' in c), 'Valor')
        col_cliente = next((c for c in col_names if 'Nome_do_cliente' in c or 'Cliente' in c), 'Nome_do_cliente')
        col_endereco = next((c for c in col_names if 'Endereco' in c), 'Endereco_do_cliente')
        col_romaneio = next((c for c in col_names if 'Lista_Entrega_Cruzada' in c), 'Lista_Entrega_Cruzada')
        
        # Formatar datas se enviadas em formato YYYY-MM-DD para DD/MM/YYYY
        def formatar_data_br(dt):
            if not dt: return ""
            if "-" in dt:
                partes = dt.split("-")
                if len(partes) == 3:
                    return f"{partes[2]}/{partes[1]}/{partes[0]}"
            return dt
            
        data_rota = formatar_data_br(dados.get("data_rota"))
        data_rec = formatar_data_br(dados.get("data_rec"))
        
        sql = f"""
            UPDATE chamados
            SET "{col_data_rota}" = ?,
                "{col_data_rec}" = ?,
                "{col_motorista}" = ?,
                "{col_placa}" = ?,
                "{col_rota}" = ?,
                "{col_regional}" = ?,
                "{col_valor}" = ?,
                "{col_cliente}" = ?,
                "{col_endereco}" = ?,
                "{col_romaneio}" = ?
            WHERE ID_do_Pedido = ?
        """
        
        cursor.execute(sql, (
            data_rota,
            data_rec,
            dados.get("motorista"),
            dados.get("placa"),
            dados.get("rota"),
            dados.get("regional"),
            dados.get("valor"),
            dados.get("cliente"),
            dados.get("endereco"),
            dados.get("romaneio"),
            pedido
        ))
        
        # Registrar no histórico todas as mudanças de campos!
        # datetime já importado no topo do arquivo
        data_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        usuario = session.get('nome', 'Administrador')
        
        valores_antigos = dict(row)
        
        def registrar_historico(campo, col_db, novo_val):
            val_ant = str(valores_antigos.get(col_db) or "").strip()
            val_nov = str(novo_val or "").strip()
            if val_ant != val_nov:
                cursor.execute("""
                    INSERT INTO historico_chamados (pedido_id, data_hora, usuario, campo, valor_antigo, valor_novo)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (pedido, data_hora, usuario, campo, val_ant if val_ant else "Vazio", val_nov if val_nov else "Vazio"))
                
        registrar_historico("Data Carregamento", col_data_rota, data_rota)
        registrar_historico("Data Reclamação", col_data_rec, data_rec)
        registrar_historico("Motorista", col_motorista, dados.get("motorista"))
        registrar_historico("Placa veículo", col_placa, dados.get("placa"))
        registrar_historico("Rota", col_rota, dados.get("rota"))
        registrar_historico("Regional", col_regional, dados.get("regional"))
        registrar_historico("Valor", col_valor, dados.get("valor"))
        registrar_historico("Nome Cliente", col_cliente, dados.get("cliente"))
        registrar_historico("Endereço Cliente", col_endereco, dados.get("endereco"))
        registrar_historico("Nº Romaneio", col_romaneio, dados.get("romaneio"))
        
        conn.commit()
        return jsonify(sucesso=True)
    except Exception as e:
        return jsonify(erro=str(e)), 500

def formatar_data_para_iso(dt_str):
    dt_str = str(dt_str).strip()
    if "/" in dt_str:
        partes = dt_str.split("/")
        if len(partes) == 3:
            d, m, y = partes
            if len(d) <= 2 and len(m) <= 2 and len(y) == 4:
                return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    return dt_str

def formatar_valor_reais(val):
    if not val:
        return ""
    try:
        val_str = str(val).replace(",", ".").strip()
        val_float = float(val_str)
        partes = f"{val_float:.2f}".split(".")
        inteiro = partes[0]
        decimal = partes[1]
        inteiro_formatado = ""
        for i, digito in enumerate(reversed(inteiro)):
            if i > 0 and i % 3 == 0:
                inteiro_formatado = "." + inteiro_formatado
            inteiro_formatado = digito + inteiro_formatado
        return f"R$ {inteiro_formatado},{decimal}"
    except (ValueError, TypeError):
        return str(val)

@app.get("/api/fluxo/autocompletar/<pedido>")
@login_required
def fluxo_autocompletar(pedido):
    pedido = str(pedido).strip()
    resultado = {
        "Motorista": "",
        "Placa_do_veiculo": "",
        "Rota": "",
        "Regional_2": "",
        "Valor": "",
        "Data_do_Carregamento": "",
        "Nome_do_cliente": "",
        "Endereco_do_cliente": ""
    }
    
    # 1. Buscar no banco SQLite
    try:
        conn = get_db()
        row = conn.execute("SELECT * FROM chamados WHERE ID_do_Pedido = ?", (pedido,)).fetchone()
        if row:
            colunas = row.keys() if hasattr(row, 'keys') else []
            col_mot = next((c for c in colunas if 'Motorista' in c), 'Motorista')
            col_placa = next((c for c in colunas if 'Placa' in c), 'Placa_do_veículo')
            col_rota = next((c for c in colunas if 'Rota' in c), 'Rota')
            col_reg = next((c for c in colunas if 'Regional' in c), 'Regional_2')
            col_val = next((c for c in colunas if 'Valor' in c), 'Valor')
            col_data = next((c for c in colunas if 'Data' in c), 'Data_do_Carregamento')
            col_cli = next((c for c in colunas if 'Nome_do_cliente' in c or 'Cliente' in c), 'Nome_do_cliente')
            col_end = next((c for c in colunas if 'Endereco_do_cliente' in c or 'Endereco' in c), 'Endereco_do_cliente')
            
            if row[col_mot]: resultado["Motorista"] = str(row[col_mot]).strip()
            if row[col_placa]: resultado["Placa_do_veiculo"] = str(row[col_placa]).strip()
            if row[col_rota]: resultado["Rota"] = str(row[col_rota]).strip()
            if row[col_reg]: resultado["Regional_2"] = str(row[col_reg]).strip()
            if row[col_val]: resultado["Valor"] = str(row[col_val]).strip()
            if row[col_data]: resultado["Data_do_Carregamento"] = formatar_data_para_iso(row[col_data])
            if col_cli in colunas and row[col_cli]: resultado["Nome_do_cliente"] = str(row[col_cli]).strip()
            if col_end in colunas and row[col_end]: resultado["Endereco_do_cliente"] = str(row[col_end]).strip()
            
            if resultado["Motorista"] and resultado["Placa_do_veiculo"] and resultado["Valor"] and resultado["Nome_do_cliente"]:
                return jsonify(resultado)
    except:
        pass

    # 2. Buscar no Gestão de Perdas -Pet Love.csv
    try:
        csv_perdas = BASE_DADOS_DIR / "Fluxo de aprovação" / "Gestão de Perdas -Pet Love.csv"
        if csv_perdas.exists():
            delimiter = ','
            with open(csv_perdas, newline='', encoding='utf-8-sig', errors='ignore') as f:
                first_line = f.readline()
                if first_line.count(';') > first_line.count(','):
                    delimiter = ';'
                f.seek(0)
                reader = csv.DictReader(f, delimiter=delimiter)
                for row in reader:
                    col_id = next((c for c in reader.fieldnames if 'ID_do_Pedido' in c or 'ID_Pedido' in c), 'ID_do_Pedido')
                    if str(row.get(col_id, "")).strip() == pedido:
                        col_mot = next((c for c in reader.fieldnames if 'Motorista' in c), 'Motorista')
                        col_placa = next((c for c in reader.fieldnames if 'Placa' in c), 'Placa do veículo')
                        col_rota = next((c for c in reader.fieldnames if 'Rota' in c), 'Rota')
                        col_reg = next((c for c in reader.fieldnames if 'Regional' in c), 'Regional_2')
                        col_val = next((c for c in reader.fieldnames if 'Valor' in c), 'Valor')
                        col_data = next((c for c in reader.fieldnames if 'Carregamento' in c or 'Data' in c), 'Data do Carregamento')
                        
                        if row.get(col_mot) and not resultado["Motorista"]: resultado["Motorista"] = str(row[col_mot]).strip()
                        if row.get(col_placa) and not resultado["Placa_do_veiculo"]: resultado["Placa_do_veiculo"] = str(row[col_placa]).strip()
                        if row.get(col_rota) and not resultado["Rota"]: resultado["Rota"] = str(row[col_rota]).strip()
                        if row.get(col_reg) and not resultado["Regional_2"]: resultado["Regional_2"] = str(row[col_reg]).strip()
                        if row.get(col_val) and not resultado["Valor"]: resultado["Valor"] = str(row[col_val]).strip()
                        if row.get(col_data) and not resultado["Data_do_Carregamento"]: resultado["Data_do_Carregamento"] = formatar_data_para_iso(row[col_data])
                        break
    except:
        pass

    # 3. Buscar na pasta Relatorio TMS
    try:
        tms_dir = BASE_DADOS_DIR / "Relatorio TMS"
        if tms_dir.exists() and tms_dir.is_dir():
            for csv_tms in tms_dir.glob("*.csv"):
                delimiter = ','
                with open(csv_tms, newline='', encoding='utf-8-sig', errors='ignore') as f:
                    first_line = f.readline()
                    if first_line.count(';') > first_line.count(','):
                        delimiter = ';'
                    f.seek(0)
                    reader = csv.DictReader(f, delimiter=delimiter)
                    if not reader.fieldnames: continue
                    for row in reader:
                        col_id = next((c for c in reader.fieldnames if 'Pedido' in c), 'Pedido')
                        if str(row.get(col_id, "")).strip() == pedido:
                            col_mot = next((c for c in reader.fieldnames if 'Motorista' in c), 'Motorista_Lista')
                            col_rota = next((c for c in reader.fieldnames if 'Rota' in c), 'Rota_Entrega')
                            col_reg = next((c for c in reader.fieldnames if 'Filial' in c), 'Filial_Entrega')
                            col_data = next((c for c in reader.fieldnames if 'Carregamento' in c), 'Data_do_Carregamento_Lista')
                            
                            if row.get(col_mot) and not resultado["Motorista"]: resultado["Motorista"] = str(row[col_mot]).strip()
                            if row.get(col_rota) and not resultado["Rota"]: resultado["Rota"] = str(row[col_rota]).strip()
                            
                            if row.get(col_reg) and not resultado["Regional_2"]: 
                                reg_raw = str(row[col_reg]).strip()
                                if "São Paulo" in reg_raw: resultado["Regional_2"] = "JM SP"
                                elif "Barueri" in reg_raw: resultado["Regional_2"] = "JM BAR"
                                elif "Santos" in reg_raw: resultado["Regional_2"] = "JM SSZ"
                                else: resultado["Regional_2"] = reg_raw
                                
                            if row.get(col_data) and not resultado["Data_do_Carregamento"]:
                                dt = str(row[col_data]).strip()
                                resultado["Data_do_Carregamento"] = formatar_data_para_iso(dt.split(" ")[0])
                            
                            col_lista = next((c for c in reader.fieldnames if 'Lista' in c and 'Entrega' in c), 'Lista_Entrega')
                            if row.get(col_lista) and not resultado.get("Lista_Entrega"):
                                resultado["Lista_Entrega"] = str(row[col_lista]).strip()
                            break
    except:
        pass

    # 4. Buscar no TMS online (API TMSLOG)
    if not all([resultado.get("Motorista"), resultado.get("Rota"), resultado.get("Regional_2"), resultado.get("Lista_Entrega")]) or not resultado.get("Valor"):
        try:
            tms_login = request.args.get("login", "")
            tms_senha = request.args.get("senha", "")
            
            if tms_login and tms_senha:
                tms_dados = buscar_dados_pedido_tms(pedido, login=tms_login, senha=tms_senha)
                if tms_dados:
                    remessa_id = tms_dados.get("id")
                    if remessa_id:
                        try:
                            from exportar_petlove import buscar_detalhes_remessa_tms
                            detalhes = buscar_detalhes_remessa_tms(remessa_id, tms_login, tms_senha)
                            if detalhes:
                                if detalhes.get("vlrnf") and not resultado.get("Valor"):
                                    resultado["Valor"] = formatar_valor_reais(detalhes["vlrnf"])
                                dest = detalhes.get("atores", [{}])[0].get("destinatario", [{}])[0]
                                if dest:
                                    if dest.get("NomeDestinatario") and not resultado.get("Nome_do_cliente"):
                                        resultado["Nome_do_cliente"] = str(dest["NomeDestinatario"]).strip()
                                    if dest.get("DsLogradouroDestinatario") and not resultado.get("Endereco_do_cliente"):
                                        logr = str(dest["DsLogradouroDestinatario"]).strip()
                                        num = str(dest.get("DsNumeroDestinatario") or "").strip()
                                        resultado["Endereco_do_cliente"] = f"{logr} {num}".strip()
                        except Exception as e:
                            print(f"Erro ao buscar detalhes da remessa: {e}")

                    lista_entrega = str(tms_dados.get("Lista_Entrega", "")).strip()
                    if lista_entrega and not resultado.get("Lista_Entrega"):
                        resultado["Lista_Entrega"] = lista_entrega
                        
                        try:
                            from exportar_petlove import consultar_lista
                            import re
                            tabelas = consultar_lista(lista_entrega, tms_login, tms_senha)
                            if tabelas and len(tabelas) > 0:
                                info_carga = str(tabelas[0][1][4])
                                
                                match_placa = re.search(r"VEÍCULO:\s*([A-Z0-9]+)", info_carga, re.IGNORECASE)
                                if match_placa and not resultado.get("Placa_do_veiculo"):
                                    resultado["Placa_do_veiculo"] = match_placa.group(1).strip()
                                    
                                match_mot = re.search(r"MOTORISTA:\s*(.*?)\s*/", info_carga, re.IGNORECASE)
                                if match_mot and not resultado.get("Motorista"):
                                    resultado["Motorista"] = match_mot.group(1).strip()
                        except Exception as e:
                            print(f"Erro ao buscar detalhes da lista: {e}")
                            
                    if tms_dados.get("Motorista_Lista") and not resultado.get("Motorista"):
                        resultado["Motorista"] = str(tms_dados["Motorista_Lista"]).strip()
                    if tms_dados.get("Rota_Entrega") and not resultado.get("Rota"):
                        resultado["Rota"] = str(tms_dados["Rota_Entrega"]).strip()
                    if tms_dados.get("Filial_Entrega") and not resultado.get("Regional_2"):
                        reg_raw = str(tms_dados["Filial_Entrega"]).strip()
                        if "São Paulo" in reg_raw: resultado["Regional_2"] = "JM SP"
                        elif "Barueri" in reg_raw: resultado["Regional_2"] = "JM BAR"
                        elif "Santos" in reg_raw: resultado["Regional_2"] = "JM SSZ"
                        else: resultado["Regional_2"] = reg_raw
                    if tms_dados.get("Data_do_Carregamento_Lista") and not resultado.get("Data_do_Carregamento"):
                        dt = str(tms_dados["Data_do_Carregamento_Lista"]).strip()
                        resultado["Data_do_Carregamento"] = formatar_data_para_iso(dt.split(" ")[0])
        except:
            pass
            
    return jsonify(resultado)

@app.post("/api/fluxo/novo")
@login_required
def fluxo_novo():
    dados = request.get_json(silent=True) or {}
    pedido = dados.get("ID_do_Pedido")
    if not pedido:
        return jsonify(erro="ID do Pedido é obrigatório."), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Verificar duplicidade
        pedido_clean = str(pedido).strip()
        cursor.execute("SELECT * FROM chamados WHERE ID_do_Pedido = ?", (pedido_clean,))
        row_existente = cursor.fetchone()
        if row_existente:
            linha = dict_from_row(row_existente)
            criador = linha.get("Criado por") or "não informado"
            status = linha.get("Status da Tratativa") or "Em Andamento"
            proc = linha.get("Procedência") or "Em Análise"
            return jsonify(erro=f"O pedido {pedido_clean} já possui uma solicitação ativa. Cadastrada por: {criador} | Status: {status} | Procedência: {proc}."), 409
            
        cursor.execute("SELECT * FROM chamados LIMIT 1")
        col_names = [description[0] for description in cursor.description]
        
        placeholders = ", ".join(["?"] * len(col_names))
        sql = f"INSERT INTO chamados VALUES ({placeholders})"
        
        valores = []
        for col in col_names:
            if col == "Status_da_Tratativa":
                valores.append("Em Andamento")
                continue
                
            key_space = col.replace("_", " ")
            if "ID do Pedido" in key_space: key_space = "ID_do_Pedido"
            if "Regional 2" in key_space: key_space = "Regional_2"
            
            valor = dados.get(key_space, dados.get(col, ""))
            valores.append(str(valor).strip())
            
        cursor.execute(sql, valores)
        conn.commit()
        return jsonify(sucesso=True)
    except Exception as e:
        return jsonify(erro=str(e)), 500

@app.get("/api/fluxo/todos")
@login_required
def fluxo_todos():
    try:
        conn = get_db()
        rows = conn.execute("SELECT * FROM chamados").fetchall()
        
        linhas = [dict_from_row(row) for row in rows]
        return jsonify(linhas)
    except Exception as e:
        return jsonify(erro=str(e)), 500

@app.get("/nova-acareacao")
@login_required
def nova_acareacao():
    return render_template("nova.html", usuario=session.get('nome'), perfil=session.get('perfil'))

@app.get("/gestao")
@login_required
def gestao():
    return render_template("gestao.html", usuario=session.get('nome'), perfil=session.get('perfil'))

@app.get("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html", usuario=session.get('nome'), perfil=session.get('perfil'))

@app.get("/apresentacao")
@login_required
def apresentacao():
    if not is_admin():
        return redirect(url_for('gestao'))
    
    # Listar e ordenar os slides numericamente
    diretorio = Path(__file__).parent / "PPT" / "Apresentação de fechamento de perdas- PET LOVE- março 2026"
    arquivos = []
    if diretorio.exists():
        import re
        def extrair_numero(nome):
            match = re.search(r"Slide(\d+)", nome, re.IGNORECASE)
            return int(match.group(1)) if match else 0
            
        arquivos = [f.name for f in diretorio.glob("Slide*.PNG")]
        arquivos.sort(key=extrair_numero)
        
    return render_template(
        "apresentacao.html",
        usuario=session.get('nome'),
        perfil=session.get('perfil'),
        slides=arquivos
    )

@app.get("/apresentacao/slides/<path:filename>")
@login_required
def obter_slide(filename):
    if session.get("perfil") != "Admin":
        return "Acesso negado", 403
    diretorio = Path(__file__).parent / "PPT" / "Apresentação de fechamento de perdas- PET LOVE- março 2026"
    return send_from_directory(diretorio, filename)

@app.get("/apresentacao/pdf")
@login_required
def gerar_apresentacao_pdf():
    if session.get("perfil") != "Admin":
        return "Acesso negado", 403
        
    diretorio = Path(__file__).parent / "PPT" / "Apresentação de fechamento de perdas- PET LOVE- março 2026"
    if not diretorio.exists():
        return "Diretório de slides não encontrado.", 404
        
    import re
    from PIL import Image
    
    arquivos = list(diretorio.glob("Slide*.PNG"))
    if not arquivos:
        arquivos = list(diretorio.glob("Slide*.png"))
        
    if not arquivos:
        return "Nenhum slide encontrado.", 404
        
    def extrair_numero(caminho):
        match = re.search(r"Slide(\d+)", caminho.name, re.IGNORECASE)
        return int(match.group(1)) if match else 0
        
    imagens = []
    for arq in arquivos:
        try:
            img = Image.open(arq)
            if img.mode != 'RGB':
                img = img.convert('RGB')
            imagens.append(img)
        except Exception as e:
            print(f"Erro ao abrir imagem {arq}: {e}")
            
    if not imagens:
        return "Erro ao processar as imagens.", 500
        
    pdf_destino = EXPORT_DIR / "apresentacao_fechamento_perdas_marco_2026.pdf"
    imagens[0].save(pdf_destino, save_all=True, append_images=imagens[1:])
    
    return send_from_directory(EXPORT_DIR, pdf_destino.name, as_attachment=True)

def sincronizar_descontos_base():
    db_path = DB_PATH
    descontos_dir = BASE_DADOS_DIR.parent / "Descontos"
    if not descontos_dir.exists():
        descontos_dir.mkdir(parents=True, exist_ok=True)
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS descontos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pedido_id TEXT UNIQUE,
            data_reclamacao TEXT,
            data_carregamento TEXT,
            motorista TEXT,
            regional TEXT,
            rota TEXT,
            placa TEXT,
            valor_desconto TEXT,
            origem_arquivo TEXT,
            data_importacao TEXT,
            observacao TEXT,
            operacao TEXT,
            motivo TEXT
        )
    ''')
    cursor.execute("PRAGMA table_info(descontos)")
    existing_cols = [c[1] for c in cursor.fetchall()]
    if 'operacao' not in existing_cols:
        cursor.execute("ALTER TABLE descontos ADD COLUMN operacao TEXT")
    if 'motivo' not in existing_cols:
        cursor.execute("ALTER TABLE descontos ADD COLUMN motivo TEXT")
    
    # Busca automática em diretórios de sincronização do OneDrive, Downloads, pasta Descontos e Histórico
    desconto_files = []
    search_dirs = [
        descontos_dir,
        descontos_dir / "Arquivo_Historico",
        Path.home() / "OneDrive - JM DISTRIBUIÇÃO",
        Path.home() / "OneDrive - JM DISTRIBUIÇÃO" / "Área de Trabalho",
        Path.home() / "Downloads",
        BASE_DADOS_DIR.parent
    ]
    
    for sdir in search_dirs:
        if sdir.exists():
            for f in sdir.glob("*.xlsx"):
                fname = f.name.lower()
                if any(w in fname for w in ['desconto', 'extravio', 'perda', 'prévia', 'previa', 'petlove', 'pet love']):
                    if f not in desconto_files and 'dashboard_acariacoes' not in fname and 'modelo' not in fname:
                        desconto_files.append(f)
                        
    import openpyxl, datetime
    now_str = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    
    total_lidos = 0
    total_importados = 0
    
    if desconto_files:
        cursor.execute("DELETE FROM descontos")
        
        for filepath in desconto_files:
            filename = filepath.name
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                sheet_names = [s for s in wb.sheetnames if any(w in s.lower() for w in ['petlove', 'pet love', 'desconto', 'extravio', 'perda', 'acareação', 'acareacao'])]
                if not sheet_names and len(wb.sheetnames) == 1:
                    sheet_names = wb.sheetnames
                    
                for sname in sheet_names:
                    ws = wb[sname]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows: continue
                    
                    header_idx = -1
                    headers = []
                    for idx, r in enumerate(rows[:6]):
                        r_clean = [str(c).strip().lower() if c is not None else '' for c in r]
                        if any(w in c for c in r_clean for w in ['pedido', 'id', 'motorista', 'rota', 'valor', 'data']):
                            header_idx = idx
                            headers = [str(c).strip() if c is not None else '' for c in r]
                            break
                    if header_idx == -1: continue
                    
                    for r in rows[header_idx+1:]:
                        if not any(r): continue
                        d_row = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
                        
                        pid = None
                        for k, v in d_row.items():
                            k_lower = k.lower()
                            if ('pedido' in k_lower or 'id' in k_lower) and v is not None and str(v).strip() != '':
                                pid = str(v).strip().split('.')[0]
                                break
                                
                        if not pid or not pid.isdigit() or len(pid) < 5: continue
                        total_lidos += 1
                        
                        dt_rec = ''
                        dt_carr = ''
                        motorista = ''
                        regional = ''
                        rota = ''
                        placa = ''
                        valor = ''
                        operacao = ''
                        motivo = ''
                        
                        for k, v in d_row.items():
                            if v is None: continue
                            k_lower = str(k).lower().strip()
                            v_str = str(v).strip()
                            if isinstance(v, datetime.datetime):
                                v_str = v.strftime('%d/%m/%Y')
                                
                            if 'reclam' in k_lower or 'solicit' in k_lower: dt_rec = v_str
                            elif 'carreg' in k_lower: dt_carr = v_str
                            elif 'motorista' in k_lower or 'driver' in k_lower: motorista = v_str
                            elif 'regional' in k_lower or 'filial' in k_lower: regional = v_str
                            elif 'rota' in k_lower: rota = v_str
                            elif 'placa' in k_lower: placa = v_str
                            elif 'valor' in k_lower or 'perda' in k_lower: valor = v_str
                            elif 'opera' in k_lower or 'cliente' in k_lower or 'conta' in k_lower: operacao = v_str
                            elif 'motivo' in k_lower: motivo = v_str

                        # Regra 1: Filtro de Operação (Manter apenas PETLOVE - LAST MILE)
                        if operacao:
                            op_clean = operacao.upper()
                            if "PETLOVE" not in op_clean:
                                continue

                        # Regra 2: Filtro de Motivo (Desconsiderar DIF FRETE, GR, MANUTENÇÃO)
                        if motivo:
                            mot_clean = motivo.upper()
                            motivos_excluidos = ["DIF FRETE", "DIFERENÇA DE FRETE", "DIF. FRETE", "GR", "MANUTENÇÃO", "MANUTENCAO"]
                            if any(m in mot_clean for m in motivos_excluidos):
                                continue

                        cursor.execute('''
                            INSERT INTO descontos (pedido_id, data_reclamacao, data_carregamento, motorista, regional, rota, placa, valor_desconto, origem_arquivo, data_importacao, operacao, motivo)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ON CONFLICT(pedido_id) DO UPDATE SET
                                data_reclamacao = coalesce(nullif(excluded.data_reclamacao, ''), descontos.data_reclamacao),
                                data_carregamento = coalesce(nullif(excluded.data_carregamento, ''), descontos.data_carregamento),
                                motorista = coalesce(nullif(excluded.motorista, ''), descontos.motorista),
                                regional = coalesce(nullif(excluded.regional, ''), descontos.regional),
                                rota = coalesce(nullif(excluded.rota, ''), descontos.rota),
                                placa = coalesce(nullif(excluded.placa, ''), descontos.placa),
                                valor_desconto = coalesce(nullif(excluded.valor_desconto, ''), descontos.valor_desconto),
                                origem_arquivo = excluded.origem_arquivo,
                                data_importacao = excluded.data_importacao,
                                operacao = coalesce(nullif(excluded.operacao, ''), descontos.operacao),
                                motivo = coalesce(nullif(excluded.motivo, ''), descontos.motivo)
                        ''', (pid, dt_rec, dt_carr, motorista, regional, rota, placa, valor, filename, now_str, operacao, motivo))
                        total_importados += 1
            except Exception as e:
                print(f"Erro ao sincronizar arquivo de desconto {filename}: {e}")
                
        conn.commit()
    conn.close()
    return total_lidos, total_importados

@app.get("/financeiro")
@login_required
def financeiro():
    if not is_admin():
        return render_template("login.html", erro="Acesso não autorizado. Apenas administradores podem acessar a área financeira."), 403
    return render_template("financeiro.html", usuario=session.get('nome'), perfil=session.get('perfil'))

@app.get("/api/financeiro/dados")
@login_required
def api_financeiro_dados():
    if not is_admin():
        return jsonify(erro="Acesso negado. Apenas administradores podem realizar esta ação."), 403
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Se a tabela descontos estiver vazia, roda sincronização inicial de fundo
        cursor.execute("SELECT COUNT(*) FROM descontos")
        if cursor.fetchone()[0] == 0:
            sincronizar_descontos_base()
        
        # 1. Casos que FORAM enviados para desconto (tabela descontos)
        cursor.execute('''
            SELECT 
                d.pedido_id,
                d.data_reclamacao,
                d.data_carregamento,
                d.motorista,
                d.regional,
                d.rota,
                d.placa,
                d.valor_desconto,
                d.origem_arquivo,
                d.data_importacao,
                d.observacao,
                d.operacao,
                d.motivo,
                c.Procedência,
                c.Status_da_Tratativa,
                c.Responsavel,
                c.Justificativa,
                c.Tratativa,
                c.Criado_por,
                c.Valor as valor_chamado,
                'Enviado para Desconto' as status_envio
            FROM descontos d
            LEFT JOIN chamados c ON d.pedido_id = c.ID_do_Pedido
            ORDER BY d.id DESC
        ''')
        rows_descontos = cursor.fetchall()
        
        # 2. Casos do fluxo de acareações que NÃO FORAM enviados para desconto
        cursor.execute('''
            SELECT 
                c.ID_do_Pedido as pedido_id,
                c.Descrição_da_Reclamação as data_reclamacao,
                c.Data_do_Carregamento as data_carregamento,
                c.Motorista as motorista,
                c.Regional_2 as regional,
                c.Rota as rota,
                c.Placa_do_veículo as placa,
                c.Valor as valor_desconto,
                'Fluxo de Acareações (Sem Desconto)' as origem_arquivo,
                '' as data_importacao,
                '' as observacao,
                'PETLOVE - LAST MILE' as operacao,
                '' as motivo,
                c.Procedência,
                c.Status_da_Tratativa,
                c.Responsavel,
                c.Justificativa,
                c.Tratativa,
                c.Criado_por,
                c.Valor as valor_chamado,
                'Não Enviado' as status_envio
            FROM chamados c
            WHERE c.ID_do_Pedido NOT IN (SELECT pedido_id FROM descontos WHERE pedido_id IS NOT NULL AND pedido_id != '')
            ORDER BY c.rowid DESC
        ''')
        rows_nao_enviados = cursor.fetchall()
        
        all_rows = list(rows_descontos) + list(rows_nao_enviados)
        
        items = []
        periodos_dict = {}
        
        qtd_enviados = 0
        valor_enviados = 0.0
        qtd_nao_enviados = 0
        valor_nao_enviados = 0.0
        
        qtd_com_chamado = 0
        qtd_sem_chamado = 0
        qtd_procedente = 0
        valor_procedente = 0.0
        qtd_nao_procedente = 0
        valor_nao_procedente = 0.0
        qtd_em_analise = 0
        valor_em_analise = 0.0
        
        def parse_num(val_str):
            if not val_str: return 0.0
            try:
                cln = str(val_str).replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
                return float(cln)
            except:
                return 0.0

        def calc_quinzena(d_row):
            dt_str = d_row.get("data_reclamacao") or d_row.get("data_carregamento") or d_row.get("Data_do_Carregamento") or ""
            if not dt_str or str(dt_str).strip() == "":
                origem = str(d_row.get("origem_arquivo") or "").lower()
                if "maio" in origem: dt_str = "15/05/2026"
                elif "abril" in origem: dt_str = "15/04/2026"
                elif "junho" in origem: dt_str = "15/06/2026"
                elif "julho" in origem: dt_str = "15/07/2026"
                else: dt_str = d_row.get("data_importacao") or "15/05/2026"

            dt_str = str(dt_str).strip()
            dt = None
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%y', '%Y/%m/%d', '%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S'):
                try:
                    dt = datetime.datetime.strptime(dt_str[:10], fmt)
                    break
                except: pass
            if not dt:
                dt = datetime.datetime(2026, 5, 15)
            
            nomes_meses = {
                1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
                5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
                9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
            }
            m_name = nomes_meses.get(dt.month, str(dt.month))
            key_month = f"{dt.year}-{dt.month:02d}"
            
            if dt.day <= 15:
                q_code = f"{key_month}-Q1"
                q_label = f"{m_name} Q1"
                sort_key = (dt.year, dt.month, 1)
            else:
                q_code = f"{key_month}-Q2"
                q_label = f"{m_name} Q2"
                sort_key = (dt.year, dt.month, 2)
                
            return q_code, q_label, sort_key

        for r in all_rows:
            d = dict(r)
            pid = str(d.get("pedido_id") or "").strip()
            val_desc_raw = str(d.get("valor_desconto") or d.get("valor_chamado") or "").strip()
            val_desc_num = parse_num(val_desc_raw)
            status_envio = d.get("status_envio") or "Não Enviado"
            
            if status_envio == "Enviado para Desconto":
                qtd_enviados += 1
                valor_enviados += val_desc_num
            else:
                qtd_nao_enviados += 1
                valor_nao_enviados += val_desc_num
            
            q_code, q_label, sort_key = calc_quinzena(d)
            if q_code and q_code not in periodos_dict:
                periodos_dict[q_code] = {"code": q_code, "label": q_label, "sort": sort_key}
            
            proc_raw = (d.get("Procedência") or "").strip()
            status_trat = (d.get("Status_da_Tratativa") or "").strip()
            resp = (d.get("Responsavel") or "").strip()
            trat = (d.get("Tratativa") or "").strip()
            
            tem_chamado = d.get("Procedência") is not None or d.get("Status_da_Tratativa") is not None or d.get("Responsavel") is not None
            
            proc_norm = "Sem Chamado"
            if tem_chamado:
                qtd_com_chamado += 1
                proc_lower = proc_raw.lower()
                esta_fin = resp != "" or trat != "" or (proc_raw != "" and "analise" not in proc_lower and "análise" not in proc_lower)
                
                if proc_raw == "Procedente" or proc_lower == "procédente":
                    proc_norm = "Procedente"
                    qtd_procedente += 1
                    valor_procedente += val_desc_num
                elif proc_raw in ["Não Procedente", "Nao Procedente", "Improcedente"]:
                    proc_norm = "Não Procedente"
                    qtd_nao_procedente += 1
                    valor_nao_procedente += val_desc_num
                elif esta_fin:
                    proc_norm = proc_raw if proc_raw else "Finalizado"
                    qtd_procedente += 1
                    valor_procedente += val_desc_num
                else:
                    proc_norm = "Em Análise"
                    qtd_em_analise += 1
                    valor_em_analise += val_desc_num
            else:
                qtd_sem_chamado += 1
                
            # Categoria de Confronto das 3 Bases (Fluxo Diário x TMSLOG x Financeiro)
            if status_envio == "Enviado para Desconto":
                if not tem_chamado:
                    status_triangulacao = "Desconto sem Acareação"
                    badge_triangulacao = "🔴 Desconto sem Chamado"
                elif proc_norm == "Procedente":
                    status_triangulacao = "Desconto Confirmado"
                    badge_triangulacao = "🟢 Confirmado & Procedente"
                elif proc_norm == "Não Procedente":
                    status_triangulacao = "Cobrança Indevida / Contestada"
                    badge_triangulacao = "🛡️ Contestado (Improcedente)"
                else:
                    status_triangulacao = "Desconto em Análise"
                    badge_triangulacao = "⏳ Desconto em Análise"
            else:
                status_triangulacao = "Pendente de Desconto"
            # Classificação de Efetivação do Desconto (Enviado para Desconto x Efetivamente Descontado)
            origem_file_lower = (d.get("origem_arquivo") or "").lower()
            is_previa_file = any(w in origem_file_lower for w in ['prévia', 'previa', 'perda', 'perdas', 'gerot_previa'])
            is_pagos_file = any(w in origem_file_lower for w in ['pago', 'pagos', 'efetivado', 'extravio', 'desconto']) or not is_previa_file

            if is_pagos_file and status_envio == "Enviado para Desconto":
                status_efetivacao = "Efetivamente Descontado"
                badge_efetivacao = "💳 Efetivamente Descontado"
                responsabilidade_custo = "Descontado Motorista"
                badge_custo = "🟢 Descontado Motorista"
            elif status_envio == "Enviado para Desconto":
                status_efetivacao = "Enviado p/ Desconto (Prévia)"
                badge_efetivacao = "📤 Enviado p/ Desconto"
                responsabilidade_custo = "Custo JM"
                badge_custo = "🏢 Custo JM"
            else:
                status_efetivacao = "Pendente de Envio"
                badge_efetivacao = "⚠️ Pendente de Envio"
                responsabilidade_custo = "Pendente Financeiro"
                badge_custo = "⏳ Pendente Financeiro"

            items.append({
                "pedido_id": pid,
                "status_envio": status_envio,
                "quinzena_code": q_code,
                "quinzena_label": q_label,
                "data_reclamacao": d.get("data_reclamacao") or "",
                "data_carregamento": d.get("data_carregamento") or "",
                "motorista": d.get("motorista") or "Não Informado",
                "regional": d.get("regional") or "",
                "rota": d.get("rota") or "",
                "placa": d.get("placa") or "",
                "valor_desconto": val_desc_raw,
                "valor_desconto_num": val_desc_num,
                "origem_arquivo": d.get("origem_arquivo") or "Fluxo de Acareações",
                "data_importacao": d.get("data_importacao") or "",
                "operacao": d.get("operacao") or "PETLOVE - LAST MILE",
                "motivo": d.get("motivo") or "",
                "status_triangulacao": status_triangulacao,
                "badge_triangulacao": badge_triangulacao,
                "responsabilidade_custo": responsabilidade_custo,
                "badge_custo": badge_custo,
                "status_efetivacao": status_efetivacao,
                "badge_efetivacao": badge_efetivacao,
                "tem_chamado": tem_chamado,
                "status_chamado": status_trat or ("Em Andamento" if tem_chamado else "Não Cadastrado"),
                "procedencia": proc_norm,
                "responsavel": d.get("Responsavel") or "",
                "justificativa": d.get("Justificativa") or "",
                "tratativa": d.get("Tratativa") or "",
                "criado_por": d.get("Criado_por") or ""
            })
            
        periodos_list = sorted(list(periodos_dict.values()), key=lambda x: x["sort"], reverse=True)
        
        # Totais do confronto das 3 bases
        val_desc_sem_chamado = sum(it["valor_desconto_num"] for it in items if it["status_triangulacao"] == "Desconto sem Acareação")
        val_desc_confirmado = sum(it["valor_desconto_num"] for it in items if it["status_triangulacao"] == "Desconto Confirmado")
        val_desc_contestavel = sum(it["valor_desconto_num"] for it in items if it["status_triangulacao"] == "Cobrança Indevida / Contestada")
        val_pend_desconto = sum(it["valor_desconto_num"] for it in items if it["status_triangulacao"] == "Pendente de Desconto")
        
        val_custo_jm = sum(it["valor_desconto_num"] for it in items if it["responsabilidade_custo"] == "Custo JM")
        qtd_custo_jm = sum(1 for it in items if it["responsabilidade_custo"] == "Custo JM")
        
        val_descontado_mot = sum(it["valor_desconto_num"] for it in items if it["responsabilidade_custo"] == "Descontado Motorista")
        qtd_descontado_mot = sum(1 for it in items if it["responsabilidade_custo"] == "Descontado Motorista")

        val_efetivamente_descontado = sum(it["valor_desconto_num"] for it in items if it["status_efetivacao"] == "Efetivamente Descontado")
        qtd_efetivamente_descontado = sum(1 for it in items if it["status_efetivacao"] == "Efetivamente Descontado")

        val_enviado_previa = sum(it["valor_desconto_num"] for it in items if it["status_efetivacao"] == "Enviado p/ Desconto (Prévia)")
        qtd_enviado_previa = sum(1 for it in items if it["status_efetivacao"] == "Enviado p/ Desconto (Prévia)")

        resumo = {
            "total_registros": len(items),
            "valor_total": valor_enviados + valor_nao_enviados,
            "qtd_enviados": qtd_enviados,
            "valor_enviados": valor_enviados,
            "qtd_nao_enviados": qtd_nao_enviados,
            "valor_nao_enviados": valor_nao_enviados,
            "qtd_com_chamado": qtd_com_chamado,
            "qtd_sem_chamado": qtd_sem_chamado,
            "qtd_procedente": qtd_procedente,
            "valor_procedente": valor_procedente,
            "qtd_nao_procedente": qtd_nao_procedente,
            "valor_nao_procedente": valor_nao_procedente,
            "qtd_em_analise": qtd_em_analise,
            "valor_em_analise": valor_em_analise,
            "qtd_desconto_sem_chamado": sum(1 for it in items if it["status_triangulacao"] == "Desconto sem Acareação"),
            "valor_desconto_sem_chamado": val_desc_sem_chamado,
            "qtd_desconto_confirmado": sum(1 for it in items if it["status_triangulacao"] == "Desconto Confirmado"),
            "valor_desconto_confirmado": val_desc_confirmado,
            "qtd_desconto_contestavel": sum(1 for it in items if it["status_triangulacao"] == "Cobrança Indevida / Contestada"),
            "valor_desconto_contestavel": val_desc_contestavel,
            "qtd_pendente_desconto": sum(1 for it in items if it["status_triangulacao"] == "Pendente de Desconto"),
            "valor_pendente_desconto": val_pend_desconto,
            "qtd_custo_jm": qtd_custo_jm,
            "valor_custo_jm": val_custo_jm,
            "qtd_descontado_motorista": qtd_descontado_mot,
            "valor_descontado_motorista": val_descontado_mot,
            "qtd_efetivamente_descontado": qtd_efetivamente_descontado,
            "valor_efetivamente_descontado": val_efetivamente_descontado,
            "qtd_enviado_previa": qtd_enviado_previa,
            "valor_enviado_previa": val_enviado_previa
        }
        
        return jsonify(registros=items, resumo=resumo, periodos=periodos_list)
    except Exception as e:
        return jsonify(erro=str(e)), 500

@app.post("/api/financeiro/sincronizar")
@login_required
def api_financeiro_sincronizar():
    if not is_admin():
        return jsonify(erro="Acesso negado. Apenas administradores podem realizar esta ação."), 403
    try:
        lidos, importados = sincronizar_descontos_base()
        return jsonify(sucesso=True, total_lidos=lidos, total_importados=importados)
    except Exception as e:
        return jsonify(erro=str(e)), 500

@app.post("/api/financeiro/upload")
@login_required
def api_financeiro_upload():
    if not is_admin():
        return jsonify(erro="Acesso negado. Apenas administradores podem realizar esta ação."), 403
    if 'arquivo' not in request.files:
        return jsonify(erro="Nenhum arquivo enviado."), 400
        
    file = request.files['arquivo']
    if file.filename == '':
        return jsonify(erro="Nenhum arquivo selecionado."), 400
        
    ext = Path(file.filename).suffix.lower()
    if ext not in ['.xlsx', '.xls', '.csv']:
        return jsonify(erro="Formato de arquivo inválido. Envie um arquivo Excel (.xlsx) ou CSV."), 400
        
    descontos_dir = BASE_DADOS_DIR.parent / "Descontos"
    descontos_dir.mkdir(parents=True, exist_ok=True)
    
    destino = descontos_dir / file.filename
    file.save(destino)
    
    try:
        lidos, importados = sincronizar_descontos_base()
        return jsonify(sucesso=True, mensagem=f"Arquivo '{file.filename}' importado com sucesso!", total_lidos=lidos, total_importados=importados)
    except Exception as e:
        return jsonify(erro=f"Erro ao processar planilha: {str(e)}"), 500

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
