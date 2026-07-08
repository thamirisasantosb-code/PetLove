import argparse
import json
import os
import sys
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE_URL = "https://petlove.tmslog.com.br"
DEFAULT_PAGE = f"{BASE_URL}/Imp_ListaPet.aspx?id=1049247"

def buscar_dados_pedido_tms(pedido, login=None, senha=None):
    if not login: login = os.getenv("TMSLOG_LOGIN")
    if not senha: senha = os.getenv("TMSLOG_SENHA")
    
    if not login or not senha:
        return None
        
    sessao = requests.Session()
    sessao.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/150.0.0.0 Safari/537.36",
        "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
    })
    
    try:
        entrada = sessao.get(f"{BASE_URL}/index.aspx", timeout=30)
        entrada.raise_for_status()
        resp_login = sessao.post(
            f"{BASE_URL}/ws.asmx/login",
            json={"usuario": login, "senha": senha},
            headers={
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "Origin": BASE_URL,
                "Referer": entrada.url,
                "X-Requested-With": "XMLHttpRequest",
            },
            timeout=30,
        )
        resp_login.raise_for_status()
        if resp_login.json().get("d", {}).get("erro") != "ok":
            return None
            
        payload = {
            "id": 16,
            "mail": "",
            "nome": "",
            "campos": "Pedido,Data do Carregamento Lista,Lista Entrega,Motorista Lista,Tipo do Veículo,Última Ocorrência,Data Última Ocorrência,Tipo da Ocorrência,Possui Comprovante,Filial Entrega,Rota Entrega,Usuario da Ocorrência",
            "filtros": f"Pedido={pedido}",
            "campocalculo": "",
            "tipogroup": "0",
            "tiposaida": "0"
        }
        
        resp_busca = sessao.post(
            f"{BASE_URL}/ws.asmx/pesquisarelatorio",
            json=payload,
            headers={
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "Origin": BASE_URL,
                "Referer": f"{BASE_URL}/rel_custom.aspx?id=16",
                "X-Requested-With": "XMLHttpRequest",
            },
            timeout=60,
        )
        resp_busca.raise_for_status()
        retorno = resp_busca.json().get("d", {})
        if retorno.get("erro"): return None
            
        json_obj = json.loads(retorno.get("json0") or "{}")
        if isinstance(json_obj, dict) and "valores" in json_obj:
            registros = json_obj["valores"]
        else:
            registros = json_obj if isinstance(json_obj, list) else [json_obj]
            
        if registros and isinstance(registros, list) and len(registros) > 0:
            return registros[0]
    except Exception:
        pass
    return None

def buscar_detalhes_remessa_tms(remessa_id, login=None, senha=None):
    if not login: login = os.getenv("TMSLOG_LOGIN")
    if not senha: senha = os.getenv("TMSLOG_SENHA")
    if not login or not senha:
        return None
        
    sessao = requests.Session()
    sessao.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/150.0.0.0 Safari/537.36",
        "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
    })
    
    try:
        entrada = sessao.get(f"{BASE_URL}/index.aspx", timeout=30)
        entrada.raise_for_status()
        resp_login = sessao.post(
            f"{BASE_URL}/ws.asmx/login",
            json={"usuario": login, "senha": senha},
            headers={
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "Origin": BASE_URL,
                "Referer": entrada.url,
                "X-Requested-With": "XMLHttpRequest",
            },
            timeout=30,
        )
        resp_login.raise_for_status()
        if resp_login.json().get("d", {}).get("erro") != "ok":
            return None
            
        resp_dados = sessao.post(
            f"{BASE_URL}/ws.asmx/DetalhesRemessa",
            json={"id": int(remessa_id)},
            headers={
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "Origin": BASE_URL,
                "Referer": f"{BASE_URL}/det_remessa.aspx?id={remessa_id}",
                "X-Requested-With": "XMLHttpRequest",
            },
            timeout=30,
        )
        resp_dados.raise_for_status()
        retorno = resp_dados.json().get("d", {})
        if retorno.get("erro"):
            return None
        
        dados = json.loads(retorno.get("json0") or "{}")
        return dados
    except Exception:
        pass
    return None


def buscar_romaneio_por_pedido(pedido, login=None, senha=None):
    dados = buscar_dados_pedido_tms(pedido, login, senha)
    if dados:
        lista = dados.get("Lista Entrega") or dados.get("Lista_Entrega")
        if lista:
            return str(lista).strip()
    return None


def consultar_lista(numero, login, senha, pedido_alvo=None):
    if not str(numero).isdigit():
        raise ValueError("O número da lista deve conter somente dígitos.")

    url_lista = f"{BASE_URL}/Imp_ListaPet.aspx?id={numero}"
    sessao = requests.Session()
    sessao.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/138 Safari/537.36",
        "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
    })

    # Reproduz a navegação real: a lista redireciona à tela de login e cria
    # a sessão que será reutilizada pelo endpoint ASMX.
    entrada = sessao.get(url_lista, timeout=30)
    entrada.raise_for_status()
    resposta = sessao.post(
        f"{BASE_URL}/ws.asmx/login",
        json={"usuario": login, "senha": senha},
        headers={
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Origin": BASE_URL,
            "Referer": entrada.url,
            "X-Requested-With": "XMLHttpRequest",
        },
        timeout=30,
    )
    resposta.raise_for_status()
    dados = resposta.json().get("d", {})
    if dados.get("erro") != "ok":
        raise RuntimeError(f"Falha no login do TMSLOG: {dados.get('erro', 'resposta inesperada')}")

    resposta_dados = sessao.post(
        f"{BASE_URL}/ws.asmx/ListaEntregaIMP",
        json={"id": int(numero)},
        headers={
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Origin": BASE_URL,
            "Referer": url_lista,
            "X-Requested-With": "XMLHttpRequest",
        },
        timeout=60,
    )
    resposta_dados.raise_for_status()
    retorno = resposta_dados.json().get("d", {})
    if retorno.get("erro"):
        raise RuntimeError(f"Falha ao carregar a lista: {retorno['erro']}")

    registros = json.loads(retorno.get("json0") or "[]")
    if isinstance(registros, dict):
        registros = [registros]
    if not registros:
        raise RuntimeError(f"A lista {numero} não possui dados.")

    primeiro = registros[0]
    informacao = BeautifulSoup(
        str(primeiro.get("dtnotafiscal", "")), "html.parser"
    ).get_text(" ", strip=True)
    resumo = [
        ["Romaneio", "Filial", "Volumes total", "Peso total", "Informação da carga"],
        [
            primeiro.get("ncarregamento", numero),
            primeiro.get("filialentrega", ""),
            primeiro.get("totalvolumescarga", primeiro.get("objetos", "")),
            primeiro.get("pesototal", ""),
            informacao,
        ],
    ]

    campos = [
        ("Pedido", "pedido"),
        ("Cliente", "nome"),
        ("Lista de Entrega", "lista_entrega"),
        ("Ordem", "orden"),
        ("Volume", "qtdvolumes"),
        ("Endereço do cliente", "enderecoCompleto"),
        ("Tipo Serviço", "tpservico"),
        ("Cidade", "cidade"),
        ("Bairro", "bairro"),
        ("Complemento", "complementodestinatario"),
    ]
    relacao = [[titulo for titulo, _ in campos]]
    for registro in registros:
        if pedido_alvo and str(registro.get("pedido", "")).strip() != str(pedido_alvo).strip():
            continue
        registro["lista_entrega"] = numero
        relacao.append([registro.get(chave, "") for _, chave in campos])
    
    # Se não encontrar o pedido alvo na lista, mas a lista existir, avisa.
    if pedido_alvo and len(relacao) == 1:
        raise RuntimeError(f"O pedido {pedido_alvo} não foi encontrado dentro da lista {numero}.")

    return [resumo, relacao]


def parse_args():
    parser = argparse.ArgumentParser(description="Exporta a tabela do TMSLOG para Excel.")
    parser.add_argument("--url", default=DEFAULT_PAGE)
    parser.add_argument("--login", default=os.getenv("TMSLOG_LOGIN"))
    parser.add_argument("--senha", default=os.getenv("TMSLOG_SENHA"))
    parser.add_argument("--saida", default="excel.xlsx")
    args = parser.parse_args()
    if not args.login or not args.senha:
        parser.error("informe --login e --senha ou TMSLOG_LOGIN e TMSLOG_SENHA")
    return args


def texto(celula):
    return " ".join(celula.stripped_strings)


def extrair_tabelas(html):
    soup = BeautifulSoup(html, "html.parser")
    tabelas = []
    for tabela in soup.find_all("table"):
        linhas = []
        for tr in tabela.find_all("tr"):
            celulas = tr.find_all(["th", "td"], recursive=False)
            if celulas:
                linhas.append([texto(celula) for celula in celulas])
        if linhas:
            tabelas.append(linhas)
    return tabelas


def salvar_excel(tabelas, destino, nomes=None):
    wb = Workbook()
    wb.remove(wb.active)
    for indice, linhas in enumerate(tabelas, 1):
        titulo = nomes[indice - 1] if nomes and indice <= len(nomes) else f"Tabela {indice}"
        ws = wb.create_sheet(titulo)
        for linha in linhas:
            ws.append(linha)
        for celula in ws[1]:
            celula.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for coluna in ws.columns:
            largura = min(max((len(str(c.value or "")) for c in coluna), default=0) + 2, 60)
            ws.column_dimensions[get_column_letter(coluna[0].column)].width = largura
    wb.save(destino)


def main():
    args = parse_args()
    sessao = requests.Session()
    sessao.headers.update({"User-Agent": "Mozilla/5.0 (TMSLOG Excel exporter)"})

    sessao.get(f"{BASE_URL}/index.aspx", timeout=30).raise_for_status()
    resposta = sessao.post(
        f"{BASE_URL}/ws.asmx/login",
        json={"usuario": args.login, "senha": args.senha},
        timeout=30,
    )
    resposta.raise_for_status()
    dados = resposta.json().get("d", {})
    if dados.get("erro") != "ok":
        raise RuntimeError(f"Falha no login do TMSLOG: {dados.get('erro', 'resposta inesperada')}")

    pagina = sessao.get(args.url, timeout=60)
    pagina.raise_for_status()
    if "login-form" in pagina.text:
        raise RuntimeError("A sessão não foi aceita; o portal retornou a tela de login.")

    tabelas = extrair_tabelas(pagina.text)
    if not tabelas:
        raise RuntimeError("Nenhuma tabela foi encontrada na página.")

    destino = Path(args.saida).resolve()
    salvar_excel(tabelas, destino)
    print(f"Arquivo criado: {destino} ({len(tabelas)} tabela(s))")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"Erro: {exc}", file=sys.stderr)
        raise SystemExit(1)
